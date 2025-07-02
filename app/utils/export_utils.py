"""
Utilidades para exportación de reportes con cabeceras institucionales estandarizadas
"""
from datetime import datetime
from flask_login import current_user
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from reportlab.platypus import Paragraph, Spacer, Table, TableStyle
from reportlab.lib.units import inch

def crear_cabecera_excel(ws, titulo_reporte, current_row=1):
    """
    Crea cabecera institucional estandarizada para Excel
    
    Args:
        ws: Worksheet de openpyxl
        titulo_reporte: Título específico del reporte
        current_row: Fila inicial (por defecto 1)
    
    Returns:
        int: Siguiente fila disponible después de la cabecera
    """
    # Estilos
    title_font = Font(name='Calibri', size=16, bold=True, color="2E5984")
    subtitle_font = Font(name='Calibri', size=12, bold=True, color="4472C4")
    info_font = Font(name='Calibri', size=10, bold=True)
    title_alignment = Alignment(horizontal="center", vertical="center")
    
    # Encabezado principal
    ws[f'A{current_row}'] = "CONSERVATORIO NACIONAL DE MÚSICA"
    ws[f'A{current_row}'].font = title_font
    ws[f'A{current_row}'].alignment = title_alignment
    ws.merge_cells(f'A{current_row}:H{current_row}')
    current_row += 1
    
    ws[f'A{current_row}'] = "Sistema de Gestión de Inventario"
    ws[f'A{current_row}'].font = subtitle_font
    ws[f'A{current_row}'].alignment = title_alignment
    ws.merge_cells(f'A{current_row}:H{current_row}')
    current_row += 1
    
    ws[f'A{current_row}'] = "Cochapata E12-56, Quito - Ecuador"
    ws[f'A{current_row}'].font = Font(name='Calibri', size=10, color="4472C4")
    ws[f'A{current_row}'].alignment = title_alignment
    ws.merge_cells(f'A{current_row}:H{current_row}')
    current_row += 3
    
    # Información del reporte
    ws[f'A{current_row}'] = "INFORMACIÓN DEL REPORTE"
    ws[f'A{current_row}'].font = subtitle_font
    ws.merge_cells(f'A{current_row}:H{current_row}')
    current_row += 1
    
    info_data = [
        ["Tipo de Reporte:", titulo_reporte],
        ["Generado por:", current_user.username if current_user.is_authenticated else 'Sistema'],
        ["Fecha:", datetime.now().strftime('%d/%m/%Y %H:%M:%S')]
    ]
    
    for label, value in info_data:
        ws[f'A{current_row}'] = label
        ws[f'A{current_row}'].font = info_font
        ws[f'B{current_row}'] = value
        current_row += 1
    
    current_row += 2
    return current_row

def crear_estilos_excel():
    """Crear estilos estandarizados para Excel"""
    return {
        'header_font': Font(name='Calibri', size=12, bold=True, color="FFFFFF"),
        'header_fill': PatternFill(start_color="2E5984", end_color="2E5984", fill_type="solid"),
        'header_alignment': Alignment(horizontal="center", vertical="center"),
        'header_border': Border(left=Side(style='thin'), right=Side(style='thin'), 
                               top=Side(style='thin'), bottom=Side(style='thin')),
        'data_font': Font(name='Calibri', size=10),
        'data_alignment': Alignment(horizontal="center", vertical="center"),
        'data_border': Border(left=Side(style='thin'), right=Side(style='thin'), 
                             top=Side(style='thin'), bottom=Side(style='thin'))
    }

def crear_tabla_detallada_excel(ws, headers, data, start_row, titulo_seccion=None):
    """
    Crear tabla detallada con formato profesional en Excel
    
    Args:
        ws: Worksheet de openpyxl
        headers: Lista de encabezados de columna
        data: Lista de listas con los datos
        start_row: Fila inicial
        titulo_seccion: Título opcional de la sección
    
    Returns:
        int: Siguiente fila disponible
    """
    estilos = crear_estilos_excel()
    current_row = start_row
    
    # Agregar título de sección si se proporciona
    if titulo_seccion:
        ws[f'A{current_row}'] = titulo_seccion
        ws[f'A{current_row}'].font = Font(name='Calibri', size=14, bold=True, color="2E5984")
        ws.merge_cells(f'A{current_row}:{chr(64 + len(headers))}{current_row}')
        current_row += 2
    
    # Crear encabezados
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=current_row, column=col, value=header)
        cell.font = estilos['header_font']
        cell.fill = estilos['header_fill']
        cell.alignment = estilos['header_alignment']
        cell.border = estilos['header_border']
    current_row += 1
    
    # Agregar datos con formato alternado
    for row_idx, row_data in enumerate(data):
        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=current_row, column=col, value=value)
            cell.font = estilos['data_font']
            cell.border = estilos['data_border']
            cell.alignment = estilos['data_alignment']
            
            # Fondo alternado para mejor legibilidad
            if row_idx % 2 == 0:
                cell.fill = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")
        current_row += 1
    
    current_row += 2  # Espacio después de la tabla
    return current_row

def crear_cabecera_pdf(titulo_reporte):
    """
    Crea cabecera institucional estandarizada para PDF
    
    Args:
        titulo_reporte: Título específico del reporte
    
    Returns:
        list: Lista de elementos para agregar al PDF
    """
    elements = []
    styles = getSampleStyleSheet()
    
    # Estilos personalizados
    title_style = ParagraphStyle(
        'TituloInstitucional',
        parent=styles['Title'],
        fontSize=18,
        fontName='Helvetica-Bold',
        textColor=colors.HexColor('#2E5984'),
        alignment=1,
        spaceAfter=10
    )
    
    subtitle_style = ParagraphStyle(
        'SubtituloInstitucional',
        parent=styles['Normal'],
        fontSize=12,
        fontName='Helvetica',
        textColor=colors.HexColor('#4472C4'),
        alignment=1,
        spaceAfter=20
    )
    
    # Encabezado institucional
    elements.append(Paragraph("CONSERVATORIO NACIONAL DE MÚSICA", title_style))
    elements.append(Paragraph("Sistema de Gestión de Inventario", subtitle_style))
    elements.append(Paragraph("Cochapata E12-56, Quito - Ecuador", subtitle_style))
    elements.append(Spacer(1, 20))
    
    # Información del reporte
    info_style = ParagraphStyle(
        'TituloSeccion',
        parent=styles['Heading2'],
        fontSize=14,
        fontName='Helvetica-Bold',
        textColor=colors.HexColor('#2E5984'),
        spaceBefore=20,
        spaceAfter=12
    )
    
    elements.append(Paragraph("INFORMACIÓN DEL REPORTE", info_style))
    
    # Tabla de información
    info_data = [
        ['Tipo de Reporte:', titulo_reporte],
        ['Generado por:', current_user.username if current_user.is_authenticated else 'Sistema'],
        ['Fecha de Generación:', datetime.now().strftime('%d/%m/%Y %H:%M:%S')]
    ]
    
    info_table = Table(info_data, colWidths=[2*inch, 3*inch])
    info_table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#F8F9FA'))
    ]))
    
    elements.append(info_table)
    elements.append(Spacer(1, 20))
    
    return elements

def crear_estilos_pdf():
    """Crear estilos estandarizados para PDF"""
    styles = getSampleStyleSheet()
    
    # Estilo para títulos de sección
    styles.add(ParagraphStyle(
        name='TituloSeccion',
        parent=styles['Heading2'],
        fontSize=14,
        fontName='Helvetica-Bold',
        textColor=colors.HexColor('#2E5984'),
        spaceBefore=20,
        spaceAfter=12
    ))
    
    return styles

def aplicar_estilo_tabla_pdf(table):
    """Aplicar estilo estandarizado a tablas PDF"""
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2E5984')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))
    return table

def ajustar_columnas_excel(ws):
    """Ajustar automáticamente el ancho de columnas en Excel"""
    from openpyxl.cell.cell import MergedCell
    
    for column in ws.columns:
        max_length = 0
        column_letter = None
        
        for cell in column:
            # Saltar celdas combinadas
            if isinstance(cell, MergedCell):
                continue
                
            # Obtener la letra de la columna de la primera celda válida
            if column_letter is None:
                column_letter = cell.column_letter
            
            try:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except Exception:
                pass
        
        # Solo ajustar si encontramos una columna válida
        if column_letter:
            adjusted_width = min(max_length + 2, 50) if max_length > 0 else 15
            ws.column_dimensions[column_letter].width = adjusted_width

def crear_tabla_detallada_pdf(headers, data, titulo_seccion=None, col_widths=None):
    """
    Crear tabla detallada con formato profesional en PDF
    
    Args:
        headers: Lista de encabezados de columna
        data: Lista de listas con los datos
        titulo_seccion: Título opcional de la sección
        col_widths: Anchos de columna personalizados
    
    Returns:
        list: Lista de elementos para agregar al PDF
    """
    from reportlab.lib.styles import getSampleStyleSheet
    
    elements = []
    styles = getSampleStyleSheet()
    
    # Agregar título de sección si se proporciona
    if titulo_seccion:
        title_style = ParagraphStyle(
            'TituloSeccionDetallada',
            parent=styles['Heading2'],
            fontSize=14,
            fontName='Helvetica-Bold',
            textColor=colors.HexColor('#2E5984'),
            spaceBefore=20,
            spaceAfter=12
        )
        elements.append(Paragraph(titulo_seccion, title_style))
    
    # Preparar datos de la tabla
    table_data = [headers] + data
    
    # Crear tabla con anchos automáticos si no se especifican
    if not col_widths:
        col_widths = [1.2*inch] * len(headers)
    
    table = Table(table_data, colWidths=col_widths)
    
    # Aplicar estilo detallado
    table.setStyle(TableStyle([
        # Encabezado
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2E5984')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        
        # Datos con fondo alternado
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        
        # Filas alternadas para mejor legibilidad
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F8F9FA')]),
    ]))
    
    elements.append(table)
    elements.append(Spacer(1, 15))
    
    return elements

def formatear_valor_moneda(valor):
    """Formatear valor como moneda"""
    try:
        return f"${float(valor):,.2f}"
    except (ValueError, TypeError):
        return "$0.00"

def formatear_fecha(fecha):
    """Formatear fecha para mostrar"""
    if hasattr(fecha, 'strftime'):
        return fecha.strftime('%d/%m/%Y')
    return str(fecha) if fecha else 'N/A'

def truncar_texto(texto, max_length=30):
    """Truncar texto para mejor visualización"""
    if not texto:
        return 'N/A'
    texto_str = str(texto)
    return texto_str[:max_length] + '...' if len(texto_str) > max_length else texto_str