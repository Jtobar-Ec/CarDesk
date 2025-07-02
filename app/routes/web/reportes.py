from flask import Blueprint, render_template, request, jsonify, make_response
from flask_login import login_required, current_user
from datetime import datetime, date
from sqlalchemy import func, extract, and_
from app.database.models import (
    MovimientoDetalle, Item, Articulo, Instrumento,
    Proveedor, Persona, Consumo, Entrada, Usuario
)
from app.database import db
from app.utils.export_utils import (
    crear_cabecera_excel, crear_estilos_excel, ajustar_columnas_excel,
    crear_cabecera_pdf, crear_estilos_pdf, aplicar_estilo_tabla_pdf,
    crear_tabla_detallada_excel, crear_tabla_detallada_pdf,
    formatear_valor_moneda, formatear_fecha, truncar_texto
)
import io
from openpyxl import Workbook
from reportlab.lib.pagesizes import A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.units import inch
from reportlab.lib import colors

bp = Blueprint('reportes', __name__)

def _mapear_tipo_movimiento(tipo_original):
    """Mapear tipos de movimiento para mostrar solo ingresos y egresos"""
    mapeo = {
        'entrada': 'Ingreso',
        'salida': 'Egreso'
    }
    return mapeo.get(tipo_original, None)

@bp.route('/')
@login_required
def index():
    """Página principal de reportes"""
    return render_template('reportes/index.html')

@bp.route('/generar', methods=['POST'])
@login_required
def generar_reporte():
    """Generar reporte según los parámetros seleccionados"""
    data = request.get_json()
    tipo_reporte = data.get('tipo_reporte')
    periodo = data.get('periodo')
    fecha_inicio = data.get('fecha_inicio')
    fecha_fin = data.get('fecha_fin')
    articulo_id = data.get('articulo_id')
    tipo_item = data.get('tipo_item')
    
    try:
        if tipo_reporte == 'movimientos':
            resultado = _generar_reporte_movimientos(periodo, fecha_inicio, fecha_fin, articulo_id, tipo_item)
        elif tipo_reporte == 'inventario_articulos':
            resultado = _generar_reporte_inventario_articulos()
        elif tipo_reporte == 'inventario_instrumentos':
            resultado = _generar_reporte_inventario_instrumentos()
        elif tipo_reporte == 'proveedores':
            resultado = _generar_reporte_proveedores(periodo, fecha_inicio, fecha_fin)
        elif tipo_reporte == 'consumos':
            resultado = _generar_reporte_consumos(periodo, fecha_inicio, fecha_fin)
        else:
            return jsonify({'error': 'Tipo de reporte no válido'}), 400
            
        return jsonify(resultado)
    except Exception as e:
        return jsonify({'error': str(e)}), 500

def _generar_reporte_movimientos(periodo, fecha_inicio, fecha_fin, articulo_id=None, tipo_item=None):
    """Generar reporte de movimientos (ingresos/egresos)"""
    query = db.session.query(
        MovimientoDetalle.m_tipo,
        func.sum(MovimientoDetalle.m_cantidad).label('total_cantidad'),
        func.sum(MovimientoDetalle.m_valorTotal).label('total_valor'),
        func.count(MovimientoDetalle.id).label('total_movimientos')
    )
    
    # Filtrar por tipo de ítem si se especifica
    if tipo_item:
        query = query.join(Item, MovimientoDetalle.i_id == Item.id).filter(Item.i_tipo == tipo_item)
    
    # Filtrar por fechas
    if periodo == 'personalizado' and fecha_inicio and fecha_fin:
        fecha_inicio = datetime.strptime(fecha_inicio, '%Y-%m-%d').date()
        fecha_fin = datetime.strptime(fecha_fin, '%Y-%m-%d').date()
        query = query.filter(MovimientoDetalle.m_fecha.between(fecha_inicio, fecha_fin))
    elif periodo == 'mes_actual':
        hoy = date.today()
        query = query.filter(
            extract('year', MovimientoDetalle.m_fecha) == hoy.year,
            extract('month', MovimientoDetalle.m_fecha) == hoy.month
        )
    elif periodo == 'año_actual':
        hoy = date.today()
        query = query.filter(extract('year', MovimientoDetalle.m_fecha) == hoy.year)
    
    # Filtrar por artículo específico
    if articulo_id:
        query = query.filter(MovimientoDetalle.i_id == articulo_id)
    
    # Agrupar por tipo de movimiento
    movimientos = query.group_by(MovimientoDetalle.m_tipo).all()
    
    # Obtener detalles de movimientos
    query_detalle = db.session.query(
        MovimientoDetalle,
        Item.i_codigo,
        Item.i_nombre,
        Proveedor.p_razonsocial,
        Persona.pe_nombre
    ).join(Item, MovimientoDetalle.i_id == Item.id)\
     .outerjoin(Entrada, MovimientoDetalle.e_id == Entrada.id)\
     .outerjoin(Proveedor, Entrada.p_id == Proveedor.id)\
     .outerjoin(Consumo, MovimientoDetalle.c_id == Consumo.id)\
     .outerjoin(Persona, Consumo.pe_id == Persona.id)
    
    # Filtrar por tipo de ítem si se especifica
    if tipo_item:
        query_detalle = query_detalle.filter(Item.i_tipo == tipo_item)
    
    # Aplicar los mismos filtros de fecha
    if periodo == 'personalizado' and fecha_inicio and fecha_fin:
        query_detalle = query_detalle.filter(MovimientoDetalle.m_fecha.between(fecha_inicio, fecha_fin))
    elif periodo == 'mes_actual':
        hoy = date.today()
        query_detalle = query_detalle.filter(
            extract('year', MovimientoDetalle.m_fecha) == hoy.year,
            extract('month', MovimientoDetalle.m_fecha) == hoy.month
        )
    elif periodo == 'año_actual':
        hoy = date.today()
        query_detalle = query_detalle.filter(extract('year', MovimientoDetalle.m_fecha) == hoy.year)
    
    if articulo_id:
        query_detalle = query_detalle.filter(MovimientoDetalle.i_id == articulo_id)
    
    detalles = query_detalle.order_by(MovimientoDetalle.m_fecha.desc()).limit(50).all()
    
    return {
        'tipo': 'movimientos',
        'resumen': [
            {
                'tipo_movimiento': _mapear_tipo_movimiento(mov.m_tipo),
                'total_cantidad': int(mov.total_cantidad),
                'total_valor': float(mov.total_valor),
                'total_movimientos': mov.total_movimientos
            } for mov in movimientos if _mapear_tipo_movimiento(mov.m_tipo) is not None
        ],
        'detalles': [
            {
                'fecha': mov.m_fecha.strftime('%Y-%m-%d'),
                'tipo': _mapear_tipo_movimiento(mov.m_tipo),
                'codigo_item': codigo,
                'nombre_item': nombre,
                'cantidad': mov.m_cantidad,
                'valor_unitario': float(mov.m_valorUnitario),
                'valor_total': float(mov.m_valorTotal),
                'proveedor': proveedor or 'N/A',
                'persona': persona or 'N/A',
                'observaciones': mov.m_observaciones or ''
            } for mov, codigo, nombre, proveedor, persona in detalles if _mapear_tipo_movimiento(mov.m_tipo) is not None
        ]
    }

def _generar_reporte_inventario_articulos():
    """Generar reporte de inventario de artículos"""
    articulos = db.session.query(
        Item, Articulo
    ).join(Articulo, Item.id == Articulo.i_id)\
     .filter(Item.i_tipo == 'articulo').all()
    
    return {
        'tipo': 'inventario_articulos',
        'articulos': [
            {
                'codigo': item.i_codigo,
                'nombre': item.i_nombre,
                'cantidad': item.i_cantidad,
                'valor_unitario': float(item.i_vUnitario),
                'valor_total': float(item.i_vTotal),
                'stock_min': articulo.a_stockMin,
                'stock_max': articulo.a_stockMax,
                'cuenta_contable': articulo.a_c_contable,
                'estado_stock': 'Bajo' if item.i_cantidad <= articulo.a_stockMin else 'Normal'
            } for item, articulo in articulos
        ]
    }

def _generar_reporte_inventario_instrumentos():
    """Generar reporte de inventario de instrumentos"""
    instrumentos = db.session.query(
        Item, Instrumento
    ).join(Instrumento, Item.id == Instrumento.i_id)\
     .filter(Item.i_tipo == 'instrumento').all()
    
    return {
        'tipo': 'inventario_instrumentos',
        'instrumentos': [
            {
                'codigo': item.i_codigo,
                'nombre': item.i_nombre,
                'marca': instrumento.i_marca,
                'modelo': instrumento.i_modelo,
                'serie': instrumento.i_serie,
                'estado': instrumento.i_estado,
                'valor_unitario': float(item.i_vUnitario),
                'valor_total': float(item.i_vTotal)
            } for item, instrumento in instrumentos
        ]
    }

def _generar_reporte_proveedores(periodo, fecha_inicio, fecha_fin):
    """Generar reporte de proveedores"""
    query = db.session.query(
        Proveedor,
        func.count(Entrada.id).label('total_entradas'),
        func.sum(MovimientoDetalle.m_valorTotal).label('total_compras')
    ).outerjoin(Entrada, Proveedor.id == Entrada.p_id)\
     .outerjoin(MovimientoDetalle, Entrada.id == MovimientoDetalle.e_id)
    
    # Filtrar por fechas
    if periodo == 'personalizado' and fecha_inicio and fecha_fin:
        fecha_inicio = datetime.strptime(fecha_inicio, '%Y-%m-%d').date()
        fecha_fin = datetime.strptime(fecha_fin, '%Y-%m-%d').date()
        query = query.filter(Entrada.e_fecha.between(fecha_inicio, fecha_fin))
    elif periodo == 'mes_actual':
        hoy = date.today()
        query = query.filter(
            extract('year', Entrada.e_fecha) == hoy.year,
            extract('month', Entrada.e_fecha) == hoy.month
        )
    elif periodo == 'año_actual':
        hoy = date.today()
        query = query.filter(extract('year', Entrada.e_fecha) == hoy.year)
    
    proveedores = query.group_by(Proveedor.id).all()
    
    return {
        'tipo': 'proveedores',
        'datos': [
            {
                'codigo': prov.p_codigo,
                'razon_social': prov.p_razonsocial,
                'ci_ruc': prov.p_ci_ruc,
                'direccion': prov.p_direccion or 'N/A',
                'telefono': prov.p_telefono or 'N/A',
                'correo': prov.p_correo or 'N/A',
                'total_entradas': entradas or 0,
                'total_compras': float(compras) if compras else 0.0
            } for prov, entradas, compras in proveedores
        ]
    }

def _generar_reporte_consumos(periodo, fecha_inicio, fecha_fin):
    """Generar reporte de consumos por personas"""
    query = db.session.query(
        Persona,
        func.count(Consumo.id).label('total_consumos'),
        func.sum(MovimientoDetalle.m_cantidad).label('total_cantidad'),
        func.sum(MovimientoDetalle.m_valorTotal).label('total_valor')
    ).outerjoin(Consumo, Persona.id == Consumo.pe_id)\
     .outerjoin(MovimientoDetalle, Consumo.id == MovimientoDetalle.c_id)
    
    # Filtrar por fechas
    if periodo == 'personalizado' and fecha_inicio and fecha_fin:
        fecha_inicio = datetime.strptime(fecha_inicio, '%Y-%m-%d').date()
        fecha_fin = datetime.strptime(fecha_fin, '%Y-%m-%d').date()
        query = query.filter(Consumo.c_fecha.between(fecha_inicio, fecha_fin))
    elif periodo == 'mes_actual':
        hoy = date.today()
        query = query.filter(
            extract('year', Consumo.c_fecha) == hoy.year,
            extract('month', Consumo.c_fecha) == hoy.month
        )
    elif periodo == 'año_actual':
        hoy = date.today()
        query = query.filter(extract('year', Consumo.c_fecha) == hoy.year)
    
    personas = query.group_by(Persona.id).all()
    
    return {
        'tipo': 'consumos',
        'datos': [
            {
                'codigo': pers.pe_codigo,
                'nombre': pers.pe_nombre,
                'total_consumos': consumos or 0,
                'total_cantidad': int(cantidad) if cantidad else 0,
                'total_valor': float(valor) if valor else 0.0
            } for pers, consumos, cantidad, valor in personas
        ]
    }

@bp.route('/articulos-select')
@login_required
def obtener_articulos_select():
    """Obtener lista de artículos para el select"""
    tipo = request.args.get('tipo', '')
    
    query = db.session.query(Item)
    if tipo == 'articulo':
        query = query.filter(Item.i_tipo == 'articulo')
    elif tipo == 'instrumento':
        query = query.filter(Item.i_tipo == 'instrumento')
    
    articulos = query.all()
    return jsonify([
        {
            'id': art.id,
            'codigo': art.i_codigo,
            'nombre': art.i_nombre
        } for art in articulos
    ])

@bp.route('/exportar/excel', methods=['POST'])
@login_required
def exportar_excel():
    """Exportar reporte a Excel con cabecera institucional estandarizada"""
    data = request.form.to_dict()
    tipo_reporte = data.get('tipo_reporte')
    periodo = data.get('periodo')
    fecha_inicio = data.get('fecha_inicio')
    fecha_fin = data.get('fecha_fin')
    articulo_id = data.get('articulo_id')
    tipo_item = data.get('tipo_item')
    
    try:
        # Generar los datos del reporte
        if tipo_reporte == 'movimientos':
            resultado = _generar_reporte_movimientos(periodo, fecha_inicio, fecha_fin, articulo_id, tipo_item)
        elif tipo_reporte == 'inventario_articulos':
            resultado = _generar_reporte_inventario_articulos()
        elif tipo_reporte == 'inventario_instrumentos':
            resultado = _generar_reporte_inventario_instrumentos()
        elif tipo_reporte == 'proveedores':
            resultado = _generar_reporte_proveedores(periodo, fecha_inicio, fecha_fin)
        elif tipo_reporte == 'consumos':
            resultado = _generar_reporte_consumos(periodo, fecha_inicio, fecha_fin)
        else:
            return jsonify({'error': 'Tipo de reporte no válido'}), 400
        
        # Crear archivo Excel
        output = io.BytesIO()
        workbook = Workbook()
        ws = workbook.active
        ws.title = f"Reporte {tipo_reporte.title()}"
        
        # Crear cabecera institucional estandarizada
        current_row = crear_cabecera_excel(
            ws,
            f"Reporte de {tipo_reporte.replace('_', ' ').title()}",
            current_user.username if current_user.is_authenticated else 'Sistema'
        )
        
        # Crear reporte completo
        _crear_reporte_completo_optimizado(ws, tipo_reporte, periodo, fecha_inicio, fecha_fin, resultado, current_row)
        
        # Ajustar columnas
        ajustar_columnas_excel(ws)
        
        # Guardar workbook
        workbook.save(output)
        output.seek(0)
        
        # Crear respuesta
        response = make_response(output.getvalue())
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        response.headers['Content-Disposition'] = f'attachment; filename=CNM_Reporte_{tipo_reporte.title()}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        
        return response
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

def _crear_reporte_completo_optimizado(ws, tipo_reporte, periodo, fecha_inicio, fecha_fin, resultado, current_row):
    """Crear reporte completo optimizado usando las utilidades centralizadas"""
    estilos = crear_estilos_excel()
    
    # Contenido específico según el tipo de reporte
    if tipo_reporte == 'movimientos':
        current_row = _agregar_seccion_movimientos(ws, resultado, current_row, estilos)
    elif tipo_reporte == 'inventario_articulos':
        current_row = _agregar_seccion_inventario_articulos(ws, resultado, current_row, estilos)
    elif tipo_reporte == 'inventario_instrumentos':
        current_row = _agregar_seccion_inventario_instrumentos(ws, resultado, current_row, estilos)
    elif tipo_reporte == 'proveedores':
        current_row = _agregar_seccion_proveedores(ws, resultado, current_row, estilos)
    elif tipo_reporte == 'consumos':
        current_row = _agregar_seccion_consumos(ws, resultado, current_row, estilos)
    
    return current_row

# Funciones auxiliares para Excel (mantenidas para compatibilidad)
def _aplicar_estilos_excel():
    """Definir estilos para Excel"""
    return {
        'header_font': Font(name='Calibri', size=12, bold=True, color="FFFFFF"),
        'header_fill': PatternFill(start_color="2E5984", end_color="2E5984", fill_type="solid"),
        'header_alignment': Alignment(horizontal="center", vertical="center"),
        'header_border': Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin')),
        'title_font': Font(name='Calibri', size=16, bold=True, color="2E5984"),
        'subtitle_font': Font(name='Calibri', size=14, bold=True, color="4472C4"),
        'title_alignment': Alignment(horizontal="center", vertical="center"),
        'data_font': Font(name='Calibri', size=10),
        'data_alignment': Alignment(horizontal="center", vertical="center"),
        'data_border': Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    }

def _crear_reporte_completo(workbook, tipo_reporte, periodo, fecha_inicio, fecha_fin, resultado):
    """Crear reporte completo en una sola hoja"""
    ws = workbook.active
    ws.title = f"Reporte {tipo_reporte.title()}"
    estilos = _aplicar_estilos_excel()
    
    current_row = 1
    
    # Encabezado principal
    ws[f'A{current_row}'] = "CONSERVATORIO NACIONAL DE MÚSICA"
    ws[f'A{current_row}'].font = estilos['title_font']
    ws[f'A{current_row}'].alignment = estilos['title_alignment']
    ws.merge_cells(f'A{current_row}:H{current_row}')
    current_row += 1
    
    ws[f'A{current_row}'] = "Sistema de Gestión de Inventario"
    ws[f'A{current_row}'].font = Font(name='Calibri', size=12, bold=True, color="4472C4")
    ws[f'A{current_row}'].alignment = Alignment(horizontal="center")
    ws.merge_cells(f'A{current_row}:H{current_row}')
    current_row += 3
    
    # Información del reporte
    ws[f'A{current_row}'] = "INFORMACIÓN DEL REPORTE"
    ws[f'A{current_row}'].font = estilos['subtitle_font']
    ws.merge_cells(f'A{current_row}:H{current_row}')
    current_row += 1
    
    info_data = [
        ["Tipo de Reporte:", tipo_reporte.replace('_', ' ').title()],
        ["Generado por:", current_user.username if current_user.is_authenticated else 'Sistema'],
        ["Fecha:", datetime.now().strftime('%d/%m/%Y %H:%M:%S')],
        ["Período:", periodo.replace('_', ' ').title() if periodo else 'N/A']
    ]
    
    for label, value in info_data:
        ws[f'A{current_row}'] = label
        ws[f'A{current_row}'].font = Font(bold=True)
        ws[f'B{current_row}'] = value
        current_row += 1
    
    current_row += 2
    
    # Contenido específico según el tipo de reporte
    if tipo_reporte == 'movimientos':
        current_row = _agregar_seccion_movimientos(ws, resultado, current_row, estilos)
    elif tipo_reporte == 'inventario_articulos':
        current_row = _agregar_seccion_inventario_articulos(ws, resultado, current_row, estilos)
    elif tipo_reporte == 'inventario_instrumentos':
        current_row = _agregar_seccion_inventario_instrumentos(ws, resultado, current_row, estilos)
    elif tipo_reporte == 'proveedores':
        current_row = _agregar_seccion_proveedores(ws, resultado, current_row, estilos)
    elif tipo_reporte == 'consumos':
        current_row = _agregar_seccion_consumos(ws, resultado, current_row, estilos)
    
    # Ajustar anchos de columnas
    for col in range(1, 9):
        ws.column_dimensions[chr(64 + col)].width = 15

def _agregar_seccion_movimientos(ws, resultado, start_row, estilos):
    """Agregar sección de movimientos con formato detallado"""
    current_row = start_row
    
    # Resumen estadístico si existe
    if resultado.get('resumen'):
        # Calcular estadísticas adicionales
        total_ingresos = sum(item.get('total_valor', 0) for item in resultado['resumen'] if item.get('tipo_movimiento') == 'Ingreso')
        total_egresos = sum(item.get('total_valor', 0) for item in resultado['resumen'] if item.get('tipo_movimiento') == 'Egreso')
        balance = total_ingresos - total_egresos
        
        # Datos del resumen con análisis
        resumen_headers = ['Tipo Movimiento', 'Cantidad', 'Valor Total', 'Movimientos', 'Porcentaje', 'Observaciones']
        resumen_data = []
        
        total_valor_general = sum(item.get('total_valor', 0) for item in resultado['resumen'])
        
        for item in resultado['resumen']:
            valor = item.get('total_valor', 0)
            porcentaje = f"{(valor/total_valor_general*100):.1f}%" if total_valor_general > 0 else "0%"
            tipo = item.get('tipo_movimiento', '')
            
            observacion = "Flujo positivo" if tipo == 'Ingreso' else "Flujo negativo"
            
            resumen_data.append([
                tipo,
                item.get('total_cantidad', 0),
                formatear_valor_moneda(valor),
                item.get('total_movimientos', 0),
                porcentaje,
                observacion
            ])
        
        # Agregar fila de balance
        resumen_data.append([
            'BALANCE GENERAL',
            'N/A',
            formatear_valor_moneda(balance),
            'N/A',
            '100%',
            'Positivo' if balance >= 0 else 'Negativo'
        ])
        
        current_row = crear_tabla_detallada_excel(ws, resumen_headers, resumen_data, current_row, "ANÁLISIS DE MOVIMIENTOS")
        current_row += 2
    
    # Detalles con análisis si existen
    if resultado.get('detalles'):
        # Preparar datos detallados con análisis
        detalle_headers = [
            'Fecha', 'Tipo', 'Código', 'Artículo', 'Cantidad', 'V. Unitario',
            'V. Total', 'Proveedor/Persona', 'Días Transcurridos', 'Observaciones'
        ]
        
        detalle_data = []
        for detalle in resultado['detalles']:
            # Calcular días transcurridos
            try:
                fecha_mov = datetime.strptime(detalle.get('fecha', ''), '%Y-%m-%d').date()
                dias_transcurridos = (datetime.now().date() - fecha_mov).days
            except:
                dias_transcurridos = 0
            
            # Determinar responsable
            responsable = detalle.get('proveedor', 'N/A') if detalle.get('tipo') == 'Ingreso' else detalle.get('persona', 'N/A')
            
            detalle_data.append([
                formatear_fecha(detalle.get('fecha', '')),
                detalle.get('tipo', ''),
                detalle.get('codigo_item', ''),
                truncar_texto(detalle.get('nombre_item', ''), 25),
                detalle.get('cantidad', 0),
                formatear_valor_moneda(detalle.get('valor_unitario', 0)),
                formatear_valor_moneda(detalle.get('valor_total', 0)),
                truncar_texto(responsable, 20),
                dias_transcurridos,
                truncar_texto(detalle.get('observaciones', 'Sin observaciones'), 25)
            ])
        
        current_row = crear_tabla_detallada_excel(ws, detalle_headers, detalle_data, current_row, "DETALLE DE MOVIMIENTOS")
    
    return current_row

def _agregar_seccion_inventario_articulos(ws, resultado, start_row, estilos):
    """Agregar sección de inventario de artículos con análisis detallado"""
    current_row = start_row
    articulos = resultado.get('articulos', [])
    
    if not articulos:
        return current_row
    
    # Calcular estadísticas del inventario
    total_articulos = len(articulos)
    valor_total_inventario = sum(art.get('valor_total', 0) for art in articulos)
    articulos_stock_bajo = sum(1 for art in articulos if art.get('estado_stock') == 'Bajo')
    cantidad_total = sum(art.get('cantidad', 0) for art in articulos)
    
    # Resumen estadístico
    resumen_headers = ['Métrica', 'Valor', 'Porcentaje', 'Estado', 'Observaciones']
    resumen_data = [
        ['Total de Artículos', total_articulos, '100%', 'Completo', 'Artículos registrados en inventario'],
        ['Valor Total Inventario', formatear_valor_moneda(valor_total_inventario), '100%', 'Valorizado', 'Valor monetario total del inventario'],
        ['Cantidad Total', cantidad_total, 'N/A', 'Contabilizado', 'Unidades totales en stock'],
        ['Artículos Stock Bajo', articulos_stock_bajo, f"{(articulos_stock_bajo/total_articulos*100):.1f}%" if total_articulos > 0 else "0%", 'Crítico' if articulos_stock_bajo > 0 else 'Normal', 'Requieren reposición urgente'],
        ['Artículos Stock Normal', total_articulos - articulos_stock_bajo, f"{((total_articulos-articulos_stock_bajo)/total_articulos*100):.1f}%" if total_articulos > 0 else "0%", 'Normal', 'Stock dentro de parámetros normales']
    ]
    
    current_row = crear_tabla_detallada_excel(ws, resumen_headers, resumen_data, current_row, "ANÁLISIS DE INVENTARIO DE ARTÍCULOS")
    current_row += 2
    
    # Detalle de artículos con análisis
    detalle_headers = [
        'Código', 'Nombre', 'Cantidad', 'Stock Mín', 'Stock Máx', 'V. Unitario',
        'V. Total', 'Estado Stock', 'Rotación Sugerida', 'Prioridad'
    ]
    
    detalle_data = []
    for art in articulos:
        cantidad = art.get('cantidad', 0)
        stock_min = art.get('stock_min', 0)
        stock_max = art.get('stock_max', 0)
        
        # Determinar rotación sugerida
        if cantidad <= stock_min:
            rotacion = "Reposición Urgente"
            prioridad = "Alta"
        elif cantidad < (stock_min + stock_max) / 2:
            rotacion = "Reposición Normal"
            prioridad = "Media"
        else:
            rotacion = "Stock Suficiente"
            prioridad = "Baja"
        
        detalle_data.append([
            art.get('codigo', ''),
            truncar_texto(art.get('nombre', ''), 30),
            cantidad,
            stock_min,
            stock_max,
            formatear_valor_moneda(art.get('valor_unitario', 0)),
            formatear_valor_moneda(art.get('valor_total', 0)),
            art.get('estado_stock', ''),
            rotacion,
            prioridad
        ])
    
    current_row = crear_tabla_detallada_excel(ws, detalle_headers, detalle_data, current_row, "DETALLE DE INVENTARIO DE ARTÍCULOS")
    
    return current_row

def _agregar_seccion_inventario_instrumentos(ws, resultado, start_row, estilos):
    """Agregar sección de inventario de instrumentos con análisis detallado"""
    current_row = start_row
    instrumentos = resultado.get('instrumentos', [])
    
    if not instrumentos:
        return current_row
    
    # Calcular estadísticas de instrumentos
    total_instrumentos = len(instrumentos)
    valor_total_instrumentos = sum(inst.get('valor_unitario', 0) for inst in instrumentos)
    
    # Análisis por estado
    estados_count = {}
    marcas_count = {}
    
    for inst in instrumentos:
        estado = inst.get('estado', 'Sin estado')
        marca = inst.get('marca', 'Sin marca')
        estados_count[estado] = estados_count.get(estado, 0) + 1
        marcas_count[marca] = marcas_count.get(marca, 0) + 1
    
    # Resumen estadístico
    resumen_headers = ['Métrica', 'Valor', 'Porcentaje', 'Observaciones']
    resumen_data = [
        ['Total de Instrumentos', total_instrumentos, '100%', 'Instrumentos registrados en inventario'],
        ['Valor Total', formatear_valor_moneda(valor_total_instrumentos), '100%', 'Valor monetario total de instrumentos'],
        ['Promedio por Instrumento', formatear_valor_moneda(valor_total_instrumentos/total_instrumentos) if total_instrumentos > 0 else '$0.00', 'N/A', 'Valor promedio por instrumento']
    ]
    
    # Agregar análisis por estado
    for estado, cantidad in estados_count.items():
        porcentaje = f"{(cantidad/total_instrumentos*100):.1f}%" if total_instrumentos > 0 else "0%"
        resumen_data.append([f'Estado "{estado}"', cantidad, porcentaje, f'Instrumentos en estado {estado.lower()}'])
    
    current_row = crear_tabla_detallada_excel(ws, resumen_headers, resumen_data, current_row, "ANÁLISIS DE INVENTARIO DE INSTRUMENTOS")
    current_row += 2
    
    # Detalle de instrumentos con análisis
    detalle_headers = [
        'Código', 'Nombre', 'Marca', 'Modelo', 'Serie', 'Estado',
        'Valor', 'Clasificación', 'Mantenimiento', 'Observaciones'
    ]
    
    detalle_data = []
    for inst in instrumentos:
        valor = inst.get('valor_unitario', 0)
        estado = inst.get('estado', '')
        
        # Clasificación por valor
        if valor > 1000:
            clasificacion = "Alto Valor"
        elif valor > 500:
            clasificacion = "Valor Medio"
        else:
            clasificacion = "Valor Básico"
        
        # Sugerencia de mantenimiento
        if estado.lower() in ['malo', 'dañado', 'reparación']:
            mantenimiento = "Urgente"
        elif estado.lower() in ['regular', 'usado']:
            mantenimiento = "Programado"
        else:
            mantenimiento = "Preventivo"
        
        # Observaciones
        observacion = f"Instrumento {estado.lower()}, requiere mantenimiento {mantenimiento.lower()}"
        
        detalle_data.append([
            inst.get('codigo', ''),
            truncar_texto(inst.get('nombre', ''), 25),
            inst.get('marca', ''),
            truncar_texto(inst.get('modelo', ''), 15),
            inst.get('serie', ''),
            estado,
            formatear_valor_moneda(valor),
            clasificacion,
            mantenimiento,
            truncar_texto(observacion, 30)
        ])
    
    current_row = crear_tabla_detallada_excel(ws, detalle_headers, detalle_data, current_row, "DETALLE DE INVENTARIO DE INSTRUMENTOS")
    
    return current_row

def _agregar_seccion_proveedores(ws, resultado, start_row, estilos):
    """Agregar sección de proveedores con análisis detallado"""
    current_row = start_row
    proveedores = resultado.get('datos', [])
    
    if not proveedores:
        return current_row
    
    # Calcular estadísticas de proveedores
    total_proveedores = len(proveedores)
    total_compras_general = sum(prov.get('total_compras', 0) for prov in proveedores)
    total_entradas_general = sum(prov.get('total_entradas', 0) for prov in proveedores)
    proveedores_activos = sum(1 for prov in proveedores if prov.get('total_entradas', 0) > 0)
    
    # Resumen estadístico
    resumen_headers = ['Métrica', 'Valor', 'Porcentaje', 'Estado', 'Observaciones']
    resumen_data = [
        ['Total de Proveedores', total_proveedores, '100%', 'Registrados', 'Proveedores en la base de datos'],
        ['Proveedores Activos', proveedores_activos, f"{(proveedores_activos/total_proveedores*100):.1f}%" if total_proveedores > 0 else "0%", 'Operativos', 'Con al menos una entrada registrada'],
        ['Proveedores Inactivos', total_proveedores - proveedores_activos, f"{((total_proveedores-proveedores_activos)/total_proveedores*100):.1f}%" if total_proveedores > 0 else "0%", 'Sin actividad', 'Sin entradas registradas'],
        ['Total Compras', formatear_valor_moneda(total_compras_general), '100%', 'Valorizado', 'Valor total de todas las compras'],
        ['Total Entradas', total_entradas_general, 'N/A', 'Contabilizado', 'Número total de entradas registradas'],
        ['Promedio por Proveedor', formatear_valor_moneda(total_compras_general/proveedores_activos) if proveedores_activos > 0 else '$0.00', 'N/A', 'Calculado', 'Valor promedio de compras por proveedor activo']
    ]
    
    current_row = crear_tabla_detallada_excel(ws, resumen_headers, resumen_data, current_row, "ANÁLISIS DE PROVEEDORES")
    current_row += 2
    
    # Detalle de proveedores con análisis
    detalle_headers = [
        'Código', 'Razón Social', 'CI/RUC', 'Teléfono', 'Entradas', 'Total Compras',
        'Promedio por Entrada', 'Clasificación', 'Estado', 'Observaciones'
    ]
    
    detalle_data = []
    for prov in proveedores:
        total_compras = prov.get('total_compras', 0)
        total_entradas = prov.get('total_entradas', 0)
        
        # Calcular promedio por entrada
        promedio_entrada = total_compras / total_entradas if total_entradas > 0 else 0
        
        # Clasificación por volumen de compras
        if total_compras > 10000:
            clasificacion = "Proveedor Principal"
        elif total_compras > 5000:
            clasificacion = "Proveedor Importante"
        elif total_compras > 1000:
            clasificacion = "Proveedor Regular"
        else:
            clasificacion = "Proveedor Ocasional"
        
        # Estado del proveedor
        if total_entradas == 0:
            estado = "Inactivo"
            observacion = "Sin actividad registrada"
        elif total_entradas >= 10:
            estado = "Muy Activo"
            observacion = f"Proveedor frecuente con {total_entradas} entradas"
        elif total_entradas >= 5:
            estado = "Activo"
            observacion = f"Proveedor regular con {total_entradas} entradas"
        else:
            estado = "Poco Activo"
            observacion = f"Proveedor esporádico con {total_entradas} entradas"
        
        detalle_data.append([
            prov.get('codigo', ''),
            truncar_texto(prov.get('razon_social', ''), 25),
            prov.get('ci_ruc', ''),
            prov.get('telefono', 'N/A'),
            total_entradas,
            formatear_valor_moneda(total_compras),
            formatear_valor_moneda(promedio_entrada),
            clasificacion,
            estado,
            truncar_texto(observacion, 30)
        ])
    
    current_row = crear_tabla_detallada_excel(ws, detalle_headers, detalle_data, current_row, "DETALLE DE PROVEEDORES")
    
    return current_row

def _agregar_seccion_consumos(ws, resultado, start_row, estilos):
    """Agregar sección de consumos con análisis detallado"""
    current_row = start_row
    consumos = resultado.get('datos', [])
    
    if not consumos:
        return current_row
    
    # Calcular estadísticas de consumos
    total_personas = len(consumos)
    total_consumos_general = sum(cons.get('total_consumos', 0) for cons in consumos)
    total_cantidad_general = sum(cons.get('total_cantidad', 0) for cons in consumos)
    total_valor_general = sum(cons.get('total_valor', 0) for cons in consumos)
    personas_activas = sum(1 for cons in consumos if cons.get('total_consumos', 0) > 0)
    
    # Resumen estadístico
    resumen_headers = ['Métrica', 'Valor', 'Porcentaje', 'Estado', 'Observaciones']
    resumen_data = [
        ['Total de Personas', total_personas, '100%', 'Registradas', 'Personas en el sistema'],
        ['Personas con Consumos', personas_activas, f"{(personas_activas/total_personas*100):.1f}%" if total_personas > 0 else "0%", 'Activas', 'Con al menos un consumo registrado'],
        ['Personas sin Consumos', total_personas - personas_activas, f"{((total_personas-personas_activas)/total_personas*100):.1f}%" if total_personas > 0 else "0%", 'Inactivas', 'Sin consumos registrados'],
        ['Total Consumos', total_consumos_general, '100%', 'Registrados', 'Número total de consumos'],
        ['Cantidad Total Consumida', total_cantidad_general, 'N/A', 'Contabilizada', 'Unidades totales consumidas'],
        ['Valor Total Consumos', formatear_valor_moneda(total_valor_general), '100%', 'Valorizado', 'Valor monetario total de consumos'],
        ['Promedio por Persona Activa', formatear_valor_moneda(total_valor_general/personas_activas) if personas_activas > 0 else '$0.00', 'N/A', 'Calculado', 'Valor promedio de consumos por persona activa']
    ]
    
    current_row = crear_tabla_detallada_excel(ws, resumen_headers, resumen_data, current_row, "ANÁLISIS DE CONSUMOS POR PERSONAL")
    current_row += 2
    
    # Detalle de consumos con análisis
    detalle_headers = [
        'Código', 'Nombre', 'Total Consumos', 'Cantidad', 'Valor Total',
        'Promedio por Consumo', 'Clasificación', 'Nivel de Actividad', 'Observaciones'
    ]
    
    detalle_data = []
    for cons in consumos:
        total_consumos = cons.get('total_consumos', 0)
        total_cantidad = cons.get('total_cantidad', 0)
        total_valor = cons.get('total_valor', 0)
        
        # Calcular promedio por consumo
        promedio_consumo = total_valor / total_consumos if total_consumos > 0 else 0
        
        # Clasificación por valor de consumos
        if total_valor > 5000:
            clasificacion = "Alto Consumidor"
        elif total_valor > 2000:
            clasificacion = "Consumidor Frecuente"
        elif total_valor > 500:
            clasificacion = "Consumidor Regular"
        elif total_valor > 0:
            clasificacion = "Consumidor Ocasional"
        else:
            clasificacion = "Sin Consumos"
        
        # Nivel de actividad
        if total_consumos == 0:
            nivel_actividad = "Inactivo"
            observacion = "Sin consumos registrados"
        elif total_consumos >= 20:
            nivel_actividad = "Muy Activo"
            observacion = f"Usuario frecuente con {total_consumos} consumos"
        elif total_consumos >= 10:
            nivel_actividad = "Activo"
            observacion = f"Usuario regular con {total_consumos} consumos"
        elif total_consumos >= 5:
            nivel_actividad = "Moderado"
            observacion = f"Usuario moderado con {total_consumos} consumos"
        else:
            nivel_actividad = "Poco Activo"
            observacion = f"Usuario esporádico con {total_consumos} consumos"
        
        detalle_data.append([
            cons.get('codigo', ''),
            truncar_texto(cons.get('nombre', ''), 25),
            total_consumos,
            total_cantidad,
            formatear_valor_moneda(total_valor),
            formatear_valor_moneda(promedio_consumo),
            clasificacion,
            nivel_actividad,
            truncar_texto(observacion, 30)
        ])
    
    current_row = crear_tabla_detallada_excel(ws, detalle_headers, detalle_data, current_row, "DETALLE DE CONSUMOS POR PERSONAL")
    
    return current_row

@bp.route('/exportar/pdf', methods=['POST'])
@login_required
def exportar_pdf():
    """Exportar reporte a PDF con cabecera institucional estandarizada"""
    data = request.form.to_dict()
    tipo_reporte = data.get('tipo_reporte')
    periodo = data.get('periodo')
    fecha_inicio = data.get('fecha_inicio')
    fecha_fin = data.get('fecha_fin')
    articulo_id = data.get('articulo_id')
    tipo_item = data.get('tipo_item')
    
    try:
        # Generar los datos del reporte
        if tipo_reporte == 'movimientos':
            resultado = _generar_reporte_movimientos(periodo, fecha_inicio, fecha_fin, articulo_id, tipo_item)
        elif tipo_reporte == 'inventario_articulos':
            resultado = _generar_reporte_inventario_articulos()
        elif tipo_reporte == 'inventario_instrumentos':
            resultado = _generar_reporte_inventario_instrumentos()
        elif tipo_reporte == 'proveedores':
            resultado = _generar_reporte_proveedores(periodo, fecha_inicio, fecha_fin)
        elif tipo_reporte == 'consumos':
            resultado = _generar_reporte_consumos(periodo, fecha_inicio, fecha_fin)
        else:
            return jsonify({'error': 'Tipo de reporte no válido'}), 400
        
        # Crear archivo PDF
        output = io.BytesIO()
        doc = SimpleDocTemplate(output, pagesize=A4)
        elements = []
        
        # Crear cabecera institucional estandarizada
        elements.extend(crear_cabecera_pdf(f"Reporte de {tipo_reporte.replace('_', ' ').title()}"))
        
        # Crear estilos
        styles = crear_estilos_pdf()
        
        # Agregar contenido según el tipo de reporte
        if tipo_reporte == 'movimientos':
            if resultado.get('resumen'):
                elements.extend(_crear_tabla_resumen_pdf(resultado['resumen'], styles))
            if resultado.get('detalles'):
                elements.extend(_crear_tabla_detalles_pdf(resultado['detalles'], styles))
        elif tipo_reporte == 'inventario_articulos':
            elements.extend(_crear_tabla_inventario_articulos_pdf(resultado['articulos'], styles))
        elif tipo_reporte == 'inventario_instrumentos':
            elements.extend(_crear_tabla_inventario_instrumentos_pdf(resultado['instrumentos'], styles))
        elif tipo_reporte == 'proveedores':
            elements.extend(_crear_tabla_proveedores_pdf(resultado['datos'], styles))
        elif tipo_reporte == 'consumos':
            elements.extend(_crear_tabla_consumos_pdf(resultado['datos'], styles))
        
        # Construir PDF
        doc.build(elements)
        output.seek(0)
        
        # Crear respuesta
        response = make_response(output.getvalue())
        response.headers['Content-Type'] = 'application/pdf'
        response.headers['Content-Disposition'] = f'attachment; filename=CNM_Reporte_{tipo_reporte.title()}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf'
        
        return response
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# Funciones auxiliares para PDF (optimizadas con utilidades centralizadas)

def _crear_tabla_resumen_pdf(resumen, styles):
    """Crear tabla de resumen para PDF con análisis detallado"""
    # Calcular estadísticas adicionales
    total_ingresos = sum(item.get('total_valor', 0) for item in resumen if item.get('tipo_movimiento') == 'Ingreso')
    total_egresos = sum(item.get('total_valor', 0) for item in resumen if item.get('tipo_movimiento') == 'Egreso')
    balance = total_ingresos - total_egresos
    
    # Datos del resumen con análisis
    resumen_headers = ['Tipo', 'Cantidad', 'Valor', 'Movimientos', 'Porcentaje']
    resumen_data = []
    
    total_valor_general = sum(item.get('total_valor', 0) for item in resumen)
    
    for item in resumen:
        valor = item.get('total_valor', 0)
        porcentaje = f"{(valor/total_valor_general*100):.1f}%" if total_valor_general > 0 else "0%"
        
        resumen_data.append([
            item.get('tipo_movimiento', ''),
            str(item.get('total_cantidad', 0)),
            formatear_valor_moneda(valor),
            str(item.get('total_movimientos', 0)),
            porcentaje
        ])
    
    # Agregar fila de balance
    resumen_data.append([
        'BALANCE',
        'N/A',
        formatear_valor_moneda(balance),
        'N/A',
        '100%'
    ])
    
    return crear_tabla_detallada_pdf(
        resumen_headers, resumen_data,
        "ANÁLISIS DE MOVIMIENTOS",
        [1.2*inch, 1*inch, 1.2*inch, 1*inch, 1*inch]
    )

def _crear_tabla_detalles_pdf(detalles, styles):
    """Crear tabla de detalles para PDF con análisis detallado"""
    # Preparar datos detallados con análisis (limitamos a 15 registros para PDF)
    detalle_headers = ['Fecha', 'Tipo', 'Código', 'Artículo', 'Cantidad', 'V. Total', 'Días']
    detalle_data = []
    
    for detalle in detalles[:15]:  # Limitar registros para PDF
        # Calcular días transcurridos
        try:
            fecha_mov = datetime.strptime(detalle.get('fecha', ''), '%Y-%m-%d').date()
            dias_transcurridos = (datetime.now().date() - fecha_mov).days
        except:
            dias_transcurridos = 0
        
        detalle_data.append([
            formatear_fecha(detalle.get('fecha', '')),
            detalle.get('tipo', ''),
            detalle.get('codigo_item', ''),
            truncar_texto(detalle.get('nombre_item', ''), 15),
            str(detalle.get('cantidad', 0)),
            formatear_valor_moneda(detalle.get('valor_total', 0)),
            str(dias_transcurridos)
        ])
    
    elements = crear_tabla_detallada_pdf(
        detalle_headers, detalle_data,
        "DETALLE DE MOVIMIENTOS",
        [0.8*inch, 0.8*inch, 0.8*inch, 1.5*inch, 0.7*inch, 1*inch, 0.5*inch]
    )
    
    if len(detalles) > 15:
        elements.append(Paragraph(f"Mostrando 15 de {len(detalles)} registros", styles['Normal']))
        elements.append(Spacer(1, 10))
    
    return elements

def _crear_tabla_inventario_articulos_pdf(articulos, styles):
    """Crear tabla de inventario de artículos para PDF con análisis detallado"""
    if not articulos:
        return []
    
    # Calcular estadísticas del inventario
    total_articulos = len(articulos)
    valor_total_inventario = sum(art.get('valor_total', 0) for art in articulos)
    articulos_stock_bajo = sum(1 for art in articulos if art.get('estado_stock') == 'Bajo')
    
    # Resumen estadístico
    resumen_headers = ['Métrica', 'Valor', 'Porcentaje', 'Observaciones']
    resumen_data = [
        ['Total Artículos', str(total_articulos), '100%', 'Artículos en inventario'],
        ['Valor Total', formatear_valor_moneda(valor_total_inventario), '100%', 'Valor monetario total'],
        ['Stock Bajo', str(articulos_stock_bajo), f"{(articulos_stock_bajo/total_articulos*100):.1f}%" if total_articulos > 0 else "0%", 'Requieren reposición'],
        ['Stock Normal', str(total_articulos - articulos_stock_bajo), f"{((total_articulos-articulos_stock_bajo)/total_articulos*100):.1f}%" if total_articulos > 0 else "0%", 'Stock adecuado']
    ]
    
    elements = crear_tabla_detallada_pdf(
        resumen_headers, resumen_data,
        "ANÁLISIS DE INVENTARIO DE ARTÍCULOS",
        [1.5*inch, 1.5*inch, 1*inch, 2*inch]
    )
    
    # Detalle de artículos con análisis
    detalle_headers = ['Código', 'Nombre', 'Cantidad', 'Stock Mín', 'V. Total', 'Estado', 'Prioridad']
    detalle_data = []
    
    for art in articulos:
        cantidad = art.get('cantidad', 0)
        stock_min = art.get('stock_min', 0)
        
        # Determinar prioridad
        if cantidad <= stock_min:
            prioridad = "Alta"
        elif cantidad < stock_min * 1.5:
            prioridad = "Media"
        else:
            prioridad = "Baja"
        
        detalle_data.append([
            art.get('codigo', ''),
            truncar_texto(art.get('nombre', ''), 20),
            str(cantidad),
            str(stock_min),
            formatear_valor_moneda(art.get('valor_total', 0)),
            art.get('estado_stock', ''),
            prioridad
        ])
    
    elements.extend(crear_tabla_detallada_pdf(
        detalle_headers, detalle_data,
        "DETALLE DE INVENTARIO DE ARTÍCULOS",
        [0.8*inch, 1.8*inch, 0.8*inch, 0.8*inch, 1*inch, 0.8*inch, 0.8*inch]
    ))
    
    return elements

def _crear_tabla_inventario_instrumentos_pdf(instrumentos, styles):
    """Crear tabla de inventario de instrumentos para PDF con análisis detallado"""
    if not instrumentos:
        return []
    
    # Calcular estadísticas de instrumentos
    total_instrumentos = len(instrumentos)
    valor_total_instrumentos = sum(inst.get('valor_unitario', 0) for inst in instrumentos)
    
    # Análisis por estado
    estados_count = {}
    for inst in instrumentos:
        estado = inst.get('estado', 'Sin estado')
        estados_count[estado] = estados_count.get(estado, 0) + 1
    
    # Resumen estadístico
    resumen_headers = ['Métrica', 'Valor', 'Porcentaje', 'Observaciones']
    resumen_data = [
        ['Total Instrumentos', str(total_instrumentos), '100%', 'Instrumentos registrados'],
        ['Valor Total', formatear_valor_moneda(valor_total_instrumentos), '100%', 'Valor monetario total'],
        ['Promedio por Instrumento', formatear_valor_moneda(valor_total_instrumentos/total_instrumentos) if total_instrumentos > 0 else '$0.00', 'N/A', 'Valor promedio']
    ]
    
    # Agregar análisis por estado
    for estado, cantidad in estados_count.items():
        porcentaje = f"{(cantidad/total_instrumentos*100):.1f}%" if total_instrumentos > 0 else "0%"
        resumen_data.append([f'Estado "{estado}"', str(cantidad), porcentaje, f'Instrumentos en {estado.lower()}'])
    
    elements = crear_tabla_detallada_pdf(
        resumen_headers, resumen_data,
        "ANÁLISIS DE INVENTARIO DE INSTRUMENTOS",
        [1.5*inch, 1.5*inch, 1*inch, 2*inch]
    )
    
    # Detalle de instrumentos con análisis
    detalle_headers = ['Código', 'Nombre', 'Marca', 'Estado', 'Valor', 'Clasificación']
    detalle_data = []
    
    for inst in instrumentos:
        valor = inst.get('valor_unitario', 0)
        
        # Clasificación por valor
        if valor > 1000:
            clasificacion = "Alto Valor"
        elif valor > 500:
            clasificacion = "Valor Medio"
        else:
            clasificacion = "Valor Básico"
        
        detalle_data.append([
            inst.get('codigo', ''),
            truncar_texto(inst.get('nombre', ''), 20),
            inst.get('marca', ''),
            inst.get('estado', ''),
            formatear_valor_moneda(valor),
            clasificacion
        ])
    
    elements.extend(crear_tabla_detallada_pdf(
        detalle_headers, detalle_data,
        "DETALLE DE INVENTARIO DE INSTRUMENTOS",
        [0.8*inch, 1.8*inch, 1*inch, 1*inch, 1*inch, 1.2*inch]
    ))
    
    return elements

def _crear_tabla_proveedores_pdf(proveedores, styles):
    """Crear tabla de proveedores para PDF con análisis detallado"""
    if not proveedores:
        return []
    
    # Calcular estadísticas de proveedores
    total_proveedores = len(proveedores)
    total_compras_general = sum(prov.get('total_compras', 0) for prov in proveedores)
    proveedores_activos = sum(1 for prov in proveedores if prov.get('total_entradas', 0) > 0)
    
    # Resumen estadístico
    resumen_headers = ['Métrica', 'Valor', 'Porcentaje', 'Observaciones']
    resumen_data = [
        ['Total Proveedores', str(total_proveedores), '100%', 'Proveedores registrados'],
        ['Proveedores Activos', str(proveedores_activos), f"{(proveedores_activos/total_proveedores*100):.1f}%" if total_proveedores > 0 else "0%", 'Con entradas registradas'],
        ['Total Compras', formatear_valor_moneda(total_compras_general), '100%', 'Valor total de compras'],
        ['Promedio por Proveedor', formatear_valor_moneda(total_compras_general/proveedores_activos) if proveedores_activos > 0 else '$0.00', 'N/A', 'Valor promedio por proveedor activo']
    ]
    
    elements = crear_tabla_detallada_pdf(
        resumen_headers, resumen_data,
        "ANÁLISIS DE PROVEEDORES",
        [1.5*inch, 1.5*inch, 1*inch, 2*inch]
    )
    
    # Detalle de proveedores con análisis
    detalle_headers = ['Código', 'Razón Social', 'CI/RUC', 'Entradas', 'Total Compras', 'Estado']
    detalle_data = []
    
    for prov in proveedores:
        total_entradas = prov.get('total_entradas', 0)
        
        # Estado del proveedor
        if total_entradas == 0:
            estado = "Inactivo"
        elif total_entradas >= 10:
            estado = "Muy Activo"
        elif total_entradas >= 5:
            estado = "Activo"
        else:
            estado = "Poco Activo"
        
        detalle_data.append([
            prov.get('codigo', ''),
            truncar_texto(prov.get('razon_social', ''), 20),
            prov.get('ci_ruc', ''),
            str(total_entradas),
            formatear_valor_moneda(prov.get('total_compras', 0)),
            estado
        ])
    
    elements.extend(crear_tabla_detallada_pdf(
        detalle_headers, detalle_data,
        "DETALLE DE PROVEEDORES",
        [0.8*inch, 1.8*inch, 1*inch, 0.8*inch, 1.2*inch, 1*inch]
    ))
    
    return elements

def _crear_tabla_consumos_pdf(consumos, styles):
    """Crear tabla de consumos para PDF con análisis detallado"""
    if not consumos:
        return []
    
    # Calcular estadísticas de consumos
    total_personas = len(consumos)
    total_valor_general = sum(cons.get('total_valor', 0) for cons in consumos)
    personas_activas = sum(1 for cons in consumos if cons.get('total_consumos', 0) > 0)
    
    # Resumen estadístico
    resumen_headers = ['Métrica', 'Valor', 'Porcentaje', 'Observaciones']
    resumen_data = [
        ['Total Personas', str(total_personas), '100%', 'Personas registradas'],
        ['Personas Activas', str(personas_activas), f"{(personas_activas/total_personas*100):.1f}%" if total_personas > 0 else "0%", 'Con consumos registrados'],
        ['Total Valor Consumos', formatear_valor_moneda(total_valor_general), '100%', 'Valor total de consumos'],
        ['Promedio por Persona', formatear_valor_moneda(total_valor_general/personas_activas) if personas_activas > 0 else '$0.00', 'N/A', 'Valor promedio por persona activa']
    ]
    
    elements = crear_tabla_detallada_pdf(
        resumen_headers, resumen_data,
        "ANÁLISIS DE CONSUMOS POR PERSONAL",
        [1.5*inch, 1.5*inch, 1*inch, 2*inch]
    )
    
    # Detalle de consumos con análisis
    detalle_headers = ['Código', 'Nombre', 'Consumos', 'Cantidad', 'Valor Total', 'Clasificación']
    detalle_data = []
    
    for cons in consumos:
        total_valor = cons.get('total_valor', 0)
        
        # Clasificación por valor de consumos
        if total_valor > 5000:
            clasificacion = "Alto Consumidor"
        elif total_valor > 2000:
            clasificacion = "Consumidor Frecuente"
        elif total_valor > 500:
            clasificacion = "Consumidor Regular"
        elif total_valor > 0:
            clasificacion = "Consumidor Ocasional"
        else:
            clasificacion = "Sin Consumos"
        
        detalle_data.append([
            cons.get('codigo', ''),
            truncar_texto(cons.get('nombre', ''), 20),
            str(cons.get('total_consumos', 0)),
            str(cons.get('total_cantidad', 0)),
            formatear_valor_moneda(total_valor),
            clasificacion
        ])
    
    elements.extend(crear_tabla_detallada_pdf(
        detalle_headers, detalle_data,
        "DETALLE DE CONSUMOS POR PERSONAL",
        [0.8*inch, 1.8*inch, 0.8*inch, 0.8*inch, 1.2*inch, 1.4*inch]
    ))
    
    return elements