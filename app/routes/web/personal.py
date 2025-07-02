from flask import Blueprint, render_template, request, redirect, url_for, flash, jsonify, make_response
from flask_login import login_required
from app.services.personal_service import PersonalService
from datetime import date, datetime
from sqlalchemy import or_, and_
from app.database.models import Persona
from app.database import db
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from app.utils.export_utils import (
    crear_cabecera_excel, crear_estilos_excel, crear_cabecera_pdf,
    crear_estilos_pdf, aplicar_estilo_tabla_pdf, ajustar_columnas_excel,
    crear_tabla_detallada_excel, crear_tabla_detallada_pdf,
    formatear_valor_moneda, formatear_fecha, truncar_texto
)

bp = Blueprint('personal', __name__)
personal_service = PersonalService()

@bp.route('/')
@login_required
def listar_personal():
    """Lista todo el personal con filtros, paginación y exportación"""
    # Parámetros de filtrado
    buscar = request.args.get('buscar', '').strip()
    cargo = request.args.get('cargo', '')
    estado = request.args.get('estado', '')
    ordenar = request.args.get('ordenar', 'nombre')
    
    # Parámetros de paginación
    page = request.args.get('page', 1, type=int)
    per_page = 20
    
    # Verificar si es exportación
    export_type = request.args.get('export')
    if export_type in ['excel', 'pdf']:
        # Para exportación, obtener todos los registros sin paginación
        query = Persona.query
        
        # Aplicar filtros
        if buscar:
            query = query.filter(
                or_(
                    Persona.pe_nombre.ilike(f'%{buscar}%'),
                    Persona.pe_codigo.ilike(f'%{buscar}%'),
                    Persona.pe_ci.ilike(f'%{buscar}%'),
                    Persona.pe_correo.ilike(f'%{buscar}%')
                )
            )
        
        if cargo:
            query = query.filter(Persona.pe_cargo == cargo)
            
        if estado:
            query = query.filter(Persona.pe_estado == estado)
        
        # Aplicar ordenamiento
        if ordenar == 'nombre':
            query = query.order_by(Persona.pe_nombre.asc())
        elif ordenar == 'codigo':
            query = query.order_by(Persona.pe_codigo.asc())
        elif ordenar == 'cargo':
            query = query.order_by(Persona.pe_cargo.asc())
        elif ordenar == 'estado':
            query = query.order_by(Persona.pe_estado.asc())
        
        personal_export = query.all()
        
        if export_type == 'excel':
            return _exportar_personal_excel(personal_export)
        elif export_type == 'pdf':
            return _exportar_personal_pdf(personal_export)
    
    # Consulta base
    query = Persona.query
    
    # Aplicar filtros
    if buscar:
        query = query.filter(
            or_(
                Persona.pe_nombre.ilike(f'%{buscar}%'),
                Persona.pe_codigo.ilike(f'%{buscar}%'),
                Persona.pe_ci.ilike(f'%{buscar}%'),
                Persona.pe_correo.ilike(f'%{buscar}%')
            )
        )
    
    if cargo:
        query = query.filter(Persona.pe_cargo == cargo)
        
    if estado:
        query = query.filter(Persona.pe_estado == estado)
    
    # Aplicar ordenamiento
    if ordenar == 'nombre':
        query = query.order_by(Persona.pe_nombre.asc())
    elif ordenar == 'codigo':
        query = query.order_by(Persona.pe_codigo.asc())
    elif ordenar == 'cargo':
        query = query.order_by(Persona.pe_cargo.asc())
    elif ordenar == 'estado':
        query = query.order_by(Persona.pe_estado.asc())
    
    # Aplicar paginación
    pagination = query.paginate(
        page=page, per_page=per_page, error_out=False
    )
    
    # Obtener opciones para filtros
    cargos_disponibles = db.session.query(Persona.pe_cargo).distinct().filter(Persona.pe_cargo.isnot(None)).all()
    cargos_disponibles = [cargo[0] for cargo in cargos_disponibles if cargo[0]]
    
    return render_template('personal/list.html',
                         personal=pagination.items,
                         pagination=pagination,
                         cargos_disponibles=cargos_disponibles)

@bp.route('/nuevo', methods=['GET', 'POST'])
@login_required
def nuevo_personal():
    """Crear una nueva persona"""
    if request.method == 'POST':
        try:
            nombre = request.form['nombre']
            apellido = request.form['apellido']
            ci = request.form['ci']
            telefono = request.form.get('telefono')
            correo = request.form.get('correo')
            direccion = request.form.get('direccion')
            cargo = request.form.get('cargo')
            estado = request.form.get('estado', 'Activo')
            
            personal_service.crear_persona(
                nombre=nombre,
                apellido=apellido,
                ci=ci,
                telefono=telefono,
                correo=correo,
                direccion=direccion,
                cargo=cargo,
                estado=estado
            )
            
            flash('Persona creada exitosamente', 'success')
            return redirect(url_for('personal.listar_personal'))
        except Exception as e:
            flash(f'Error al crear persona: {str(e)}', 'error')
    
    # Obtener opciones para los selects
    cargos = personal_service.obtener_cargos_disponibles()
    estados = personal_service.obtener_estados_disponibles()
    return render_template('personal/form.html', cargos=cargos, estados=estados)

@bp.route('/<int:persona_id>')
@login_required
def detalle_personal(persona_id):
    """Ver detalles de una persona"""
    persona = personal_service.obtener_por_id(persona_id)
    if not persona:
        flash('Persona no encontrada', 'error')
        return redirect(url_for('personal.listar_personal'))
    
    # Obtener historial de consumos
    consumos = personal_service.obtener_historial_consumos(persona_id)
    
    # Verificar si es exportación
    export_type = request.args.get('export')
    if export_type in ['excel', 'pdf']:
        if export_type == 'excel':
            return _exportar_detalle_personal_excel(persona, consumos)
        elif export_type == 'pdf':
            return _exportar_detalle_personal_pdf(persona, consumos)
    
    today = date.today().strftime('%Y-%m-%d')
    return render_template('personal/detail.html', persona=persona, consumos=consumos, today=today)

@bp.route('/<int:persona_id>/editar', methods=['GET', 'POST'])
@login_required
def editar_personal(persona_id):
    """Editar una persona"""
    persona = personal_service.obtener_por_id(persona_id)
    if not persona:
        flash('Persona no encontrada', 'error')
        return redirect(url_for('personal.listar_personal'))
    
    if request.method == 'POST':
        try:
            # Obtener todos los campos del formulario
            datos = {
                'nombre': request.form['nombre'],
                'apellido': request.form.get('apellido', ''),
                'ci': request.form.get('ci', ''),
                'telefono': request.form.get('telefono', ''),
                'correo': request.form.get('correo', ''),
                'direccion': request.form.get('direccion', ''),
                'cargo': request.form.get('cargo', ''),
                'estado': request.form.get('estado', 'Activo')
            }
            
            personal_service.actualizar_persona(persona_id, **datos)
            flash('Persona actualizada exitosamente', 'success')
            return redirect(url_for('personal.detalle_personal', persona_id=persona_id))
        except Exception as e:
            flash(f'Error al actualizar persona: {str(e)}', 'error')
    
    # Obtener opciones para los selects
    cargos = personal_service.obtener_cargos_disponibles()
    estados = personal_service.obtener_estados_disponibles()
    return render_template('personal/form.html', persona=persona, cargos=cargos, estados=estados)

@bp.route('/<int:persona_id>/eliminar', methods=['POST'])
@login_required
def eliminar_personal(persona_id):
    """Eliminar una persona"""
    try:
        if personal_service.eliminar_persona(persona_id):
            flash('Persona eliminada exitosamente', 'success')
        else:
            flash('No se pudo eliminar la persona', 'error')
    except Exception as e:
        flash(f'Error al eliminar persona: {str(e)}', 'error')
    
    return redirect(url_for('personal.listar_personal'))

@bp.route('/<int:persona_id>/cambiar-estado', methods=['POST'])
@login_required
def cambiar_estado(persona_id):
    """Cambiar el estado de una persona"""
    try:
        nuevo_estado = request.form['estado']
        personal_service.cambiar_estado(persona_id, nuevo_estado)
        flash(f'Estado cambiado a {nuevo_estado} exitosamente', 'success')
    except Exception as e:
        flash(f'Error al cambiar estado: {str(e)}', 'error')
    
    return redirect(request.referrer or url_for('personal.listar_personal'))

@bp.route('/buscar')
@login_required
def buscar_personal():
    """Buscar personal por nombre, apellido o código"""
    termino = request.args.get('termino', '').strip()
    personal = []
    
    # Solo buscar si hay un término de búsqueda válido
    if termino and len(termino) >= 2:
        personal = personal_service.buscar_por_nombre(termino)
    elif termino and len(termino) < 2:
        # Si el término es muy corto, mostrar mensaje de error
        flash('El término de búsqueda debe tener al menos 2 caracteres', 'warning')
    
    return render_template('personal/search.html', personal=personal, termino=termino)

@bp.route('/activos')
@login_required
def personal_activo():
    """Lista solo el personal activo"""
    personal = personal_service.obtener_activos()
    return render_template('personal/list.html', personal=personal, titulo="Personal Activo")

@bp.route('/por-cargo/<cargo>')
@login_required
def personal_por_cargo(cargo):
    """Lista personal por cargo"""
    personal = personal_service.obtener_por_cargo(cargo)
    return render_template('personal/list.html', personal=personal, titulo=f"Personal - {cargo}")

@bp.route('/api/activos')
@login_required
def api_personal_activo():
    """API endpoint para obtener personal activo en formato JSON"""
    try:
        personal = personal_service.obtener_activos()
        personal_data = []
        
        for persona in personal:
            personal_data.append({
                'id': persona.id,
                'nombre': persona.pe_nombre,
                'apellido': persona.pe_apellido or '',
                'ci': persona.pe_ci or '',
                'telefono': persona.pe_telefono or '',
                'correo': persona.pe_correo or '',
                'direccion': persona.pe_direccion or '',
                'cargo': persona.pe_cargo or '',
                'estado': persona.pe_estado
            })
        
        return jsonify({
            'success': True,
            'personal': personal_data,
            'total': len(personal_data)
        })
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e),
            'personal': []
        }), 500

def _exportar_personal_excel(personal):
    """Exporta personal a Excel con formato detallado y análisis estadístico"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Personal Detallado"
    
    # Configurar orientación horizontal
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    
    # Crear cabecera institucional
    current_row = crear_cabecera_excel(ws, "Reporte Detallado de Personal")
    
    # Preparar datos detallados con estadísticas
    headers = [
        'Código', 'Nombre Completo', 'Cédula de Identidad', 'Teléfono',
        'Correo Electrónico', 'Dirección', 'Cargo', 'Estado',
        'Fecha Registro', 'Antigüedad (días)', 'Perfil Completo'
    ]
    
    data = []
    personal_por_cargo = {}
    personal_por_estado = {}
    total_activos = 0
    total_inactivos = 0
    
    for persona in personal:
        # Calcular antigüedad
        if hasattr(persona, 'created_at') and persona.created_at:
            fecha_registro = persona.created_at
            antiguedad = (datetime.now().date() - fecha_registro.date()).days if isinstance(fecha_registro, datetime) else (datetime.now().date() - fecha_registro).days
        else:
            fecha_registro = None
            antiguedad = 0
        
        # Determinar si el perfil está completo
        campos_requeridos = [persona.pe_nombre, persona.pe_ci, persona.pe_telefono, persona.pe_correo, persona.pe_cargo]
        perfil_completo = "Completo" if all(campo for campo in campos_requeridos) else "Incompleto"
        
        # Estadísticas
        cargo = persona.pe_cargo or 'Sin cargo'
        estado = persona.pe_estado or 'Sin estado'
        
        personal_por_cargo[cargo] = personal_por_cargo.get(cargo, 0) + 1
        personal_por_estado[estado] = personal_por_estado.get(estado, 0) + 1
        
        if estado.lower() == 'activo':
            total_activos += 1
        else:
            total_inactivos += 1
        
        data.append([
            persona.pe_codigo or 'N/A',
            f"{persona.pe_nombre or ''} {persona.pe_apellido or ''}".strip(),
            persona.pe_ci or 'N/A',
            persona.pe_telefono or 'N/A',
            persona.pe_correo or 'N/A',
            truncar_texto(persona.pe_direccion or 'N/A', 30),
            cargo,
            estado,
            formatear_fecha(fecha_registro) if fecha_registro else 'N/A',
            antiguedad,
            perfil_completo
        ])
    
    # Crear tabla detallada
    current_row = crear_tabla_detallada_excel(ws, headers, data, current_row, "DIRECTORIO DETALLADO DE PERSONAL")
    
    # Agregar análisis estadístico completo
    current_row += 2
    ws[f'A{current_row}'] = "ANÁLISIS ESTADÍSTICO DEL PERSONAL"
    ws[f'A{current_row}'].font = Font(name='Calibri', size=14, bold=True, color="2E5984")
    ws.merge_cells(f'A{current_row}:F{current_row}')
    current_row += 2
    
    resumen_headers = ['Métrica', 'Cantidad', 'Porcentaje', 'Observaciones']
    resumen_data = [
        ['Total de Personal', len(personal), '100%', 'Personal registrado en el sistema'],
        ['Personal Activo', total_activos, f"{(total_activos/len(personal)*100):.1f}%" if personal else "0%", 'Personal en estado activo'],
        ['Personal Inactivo', total_inactivos, f"{(total_inactivos/len(personal)*100):.1f}%" if personal else "0%", 'Personal en estado inactivo'],
        ['Perfiles Completos', sum(1 for row in data if row[10] == 'Completo'), f"{(sum(1 for row in data if row[10] == 'Completo')/len(personal)*100):.1f}%" if personal else "0%", 'Personal con información completa'],
        ['Perfiles Incompletos', sum(1 for row in data if row[10] == 'Incompleto'), f"{(sum(1 for row in data if row[10] == 'Incompleto')/len(personal)*100):.1f}%" if personal else "0%", 'Personal con información faltante']
    ]
    
    # Agregar desglose por cargo
    for cargo, cantidad in personal_por_cargo.items():
        porcentaje = f"{(cantidad/len(personal)*100):.1f}%" if personal else "0%"
        resumen_data.append([f'Personal en "{cargo}"', cantidad, porcentaje, f'Personas con cargo: {cargo}'])
    
    current_row = crear_tabla_detallada_excel(ws, resumen_headers, resumen_data, current_row)
    
    # Agregar ranking de cargos más comunes
    if personal_por_cargo:
        current_row += 2
        ws[f'A{current_row}'] = "DISTRIBUCIÓN POR CARGOS"
        ws[f'A{current_row}'].font = Font(name='Calibri', size=12, bold=True, color="2E5984")
        ws.merge_cells(f'A{current_row}:D{current_row}')
        current_row += 2
        
        cargo_headers = ['Cargo', 'Cantidad de Personal', 'Porcentaje', 'Observaciones']
        cargo_data = []
        
        # Ordenar por cantidad de personal
        top_cargos = sorted(personal_por_cargo.items(), key=lambda x: x[1], reverse=True)
        
        for cargo, cantidad in top_cargos:
            porcentaje = f"{(cantidad/len(personal)*100):.1f}%" if personal else "0%"
            cargo_data.append([
                cargo,
                cantidad,
                porcentaje,
                'Cargo principal' if cantidad == max(personal_por_cargo.values()) else 'Cargo secundario'
            ])
        
        current_row = crear_tabla_detallada_excel(ws, cargo_headers, cargo_data, current_row)
    
    # Ajustar ancho de columnas
    ajustar_columnas_excel(ws)
    
    # Crear respuesta
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    response = make_response(output.read())
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Content-Disposition'] = f'attachment; filename=CNM_Personal_Detallado_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    
    return response

def _exportar_personal_pdf(personal):
    """Exporta personal a PDF con formato detallado y análisis estadístico"""
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4),
                          rightMargin=0.5*inch, leftMargin=0.5*inch,
                          topMargin=0.5*inch, bottomMargin=0.5*inch)
    
    elements = []
    
    # Crear cabecera institucional
    elements.extend(crear_cabecera_pdf("Reporte Detallado de Personal"))
    
    # Calcular estadísticas
    total_activos = sum(1 for p in personal if p.pe_estado and p.pe_estado.lower() == 'activo')
    total_inactivos = len(personal) - total_activos
    personal_por_cargo = {}
    perfiles_completos = 0
    
    for persona in personal:
        cargo = persona.pe_cargo or 'Sin cargo'
        personal_por_cargo[cargo] = personal_por_cargo.get(cargo, 0) + 1
        
        # Verificar perfil completo
        campos_requeridos = [persona.pe_nombre, persona.pe_ci, persona.pe_telefono, persona.pe_correo, persona.pe_cargo]
        if all(campo for campo in campos_requeridos):
            perfiles_completos += 1
    
    # Resumen estadístico
    resumen_headers = ['Métrica', 'Cantidad', 'Porcentaje', 'Observaciones']
    resumen_data = [
        ['Total Personal', str(len(personal)), '100%', 'Personal registrado'],
        ['Personal Activo', str(total_activos), f"{(total_activos/len(personal)*100):.1f}%" if personal else "0%", 'En estado activo'],
        ['Personal Inactivo', str(total_inactivos), f"{(total_inactivos/len(personal)*100):.1f}%" if personal else "0%", 'En estado inactivo'],
        ['Perfiles Completos', str(perfiles_completos), f"{(perfiles_completos/len(personal)*100):.1f}%" if personal else "0%", 'Información completa'],
        ['Cargos Diferentes', str(len(personal_por_cargo)), 'N/A', 'Diversidad de cargos']
    ]
    
    elements.extend(crear_tabla_detallada_pdf(
        resumen_headers, resumen_data,
        "RESUMEN ESTADÍSTICO DEL PERSONAL",
        [1.5*inch, 1*inch, 1*inch, 2*inch]
    ))
    
    # Datos detallados de la tabla
    data_headers = ['Código', 'Nombre', 'CI', 'Teléfono', 'Correo', 'Cargo', 'Estado']
    data = []
    
    for persona in personal:
        data.append([
            persona.pe_codigo or 'N/A',
            truncar_texto(f"{persona.pe_nombre or ''} {persona.pe_apellido or ''}".strip(), 20),
            persona.pe_ci or 'N/A',
            persona.pe_telefono or 'N/A',
            truncar_texto(persona.pe_correo or 'N/A', 18),
            truncar_texto(persona.pe_cargo or 'Sin cargo', 15),
            persona.pe_estado or 'N/A'
        ])
    
    # Crear tabla detallada de personal
    elements.extend(crear_tabla_detallada_pdf(
        data_headers, data,
        "DIRECTORIO DETALLADO DE PERSONAL",
        [0.8*inch, 1.8*inch, 0.8*inch, 1*inch, 1.5*inch, 1.2*inch, 0.8*inch]
    ))
    
    # Construir PDF
    doc.build(elements)
    buffer.seek(0)
    
    response = make_response(buffer.read())
    response.headers['Content-Type'] = 'application/pdf'
    response.headers['Content-Disposition'] = f'attachment; filename=CNM_Personal_Detallado_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf'
    
    return response

def _exportar_detalle_personal_excel(persona, consumos):
    """Exporta detalle de personal a Excel con formato detallado y análisis estadístico"""
    wb = Workbook()
    ws = wb.active
    ws.title = f"Detalle_{persona.pe_nombre}"
    
    # Configurar orientación horizontal
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    
    # Crear cabecera institucional
    current_row = crear_cabecera_excel(ws, f"Detalle de Personal - {persona.pe_nombre}")
    
    # Información personal detallada
    info_headers = ['Campo', 'Valor', 'Estado', 'Observaciones']
    info_data = [
        ['Código', persona.pe_codigo or 'N/A', 'Asignado' if persona.pe_codigo else 'Sin asignar', 'Identificador único'],
        ['Nombre Completo', f"{persona.pe_nombre or ''} {persona.pe_apellido or ''}".strip(), 'Completo' if persona.pe_nombre else 'Incompleto', 'Nombre y apellido'],
        ['Cédula de Identidad', persona.pe_ci or 'N/A', 'Registrada' if persona.pe_ci else 'Faltante', 'Documento de identificación'],
        ['Teléfono', persona.pe_telefono or 'N/A', 'Registrado' if persona.pe_telefono else 'Faltante', 'Contacto telefónico'],
        ['Correo Electrónico', persona.pe_correo or 'N/A', 'Registrado' if persona.pe_correo else 'Faltante', 'Contacto por email'],
        ['Dirección', truncar_texto(persona.pe_direccion or 'N/A', 40), 'Registrada' if persona.pe_direccion else 'Faltante', 'Dirección de residencia'],
        ['Cargo', persona.pe_cargo or 'N/A', 'Asignado' if persona.pe_cargo else 'Sin asignar', 'Posición en la institución'],
        ['Estado', persona.pe_estado or 'N/A', persona.pe_estado or 'Sin definir', 'Estado actual del personal']
    ]
    
    # Crear tabla detallada de información personal
    current_row = crear_tabla_detallada_excel(ws, info_headers, info_data, current_row, "INFORMACIÓN PERSONAL DETALLADA")
    
    # Análisis de asignaciones si existen
    if consumos:
        current_row += 2
        
        # Preparar datos de asignaciones con análisis
        asig_headers = [
            'Fecha', 'Artículo', 'Código', 'Cantidad', 'Valor Unitario', 'Valor Total',
            'Estado', 'Días Transcurridos', 'Observaciones', 'Clasificación'
        ]
        
        asig_data = []
        valor_total_asignaciones = 0
        asignaciones_por_estado = {}
        cantidad_total = 0
        
        for consumo in consumos:
            # Calcular días transcurridos
            if consumo.c_fecha:
                if isinstance(consumo.c_fecha, date):
                    fecha_asignacion = consumo.c_fecha
                else:
                    fecha_asignacion = consumo.c_fecha.date()
                dias_transcurridos = (datetime.now().date() - fecha_asignacion).days
            else:
                dias_transcurridos = 0
            
            # Clasificar asignación
            if dias_transcurridos > 365:
                clasificacion = "Asignación antigua"
            elif dias_transcurridos > 90:
                clasificacion = "Asignación reciente"
            else:
                clasificacion = "Asignación nueva"
            
            # Estadísticas
            valor_total_asignaciones += float(consumo.c_valorTotal) if consumo.c_valorTotal else 0
            cantidad_total += consumo.c_cantidad or 0
            estado = consumo.c_estado or 'Sin estado'
            asignaciones_por_estado[estado] = asignaciones_por_estado.get(estado, 0) + 1
            
            asig_data.append([
                formatear_fecha(consumo.c_fecha) if consumo.c_fecha else 'N/A',
                consumo.item.i_nombre if consumo.item else 'N/A',
                consumo.item.i_codigo if consumo.item else 'N/A',
                consumo.c_cantidad or 0,
                formatear_valor_moneda(consumo.c_valorUnitario) if consumo.c_valorUnitario else '$0.00',
                formatear_valor_moneda(consumo.c_valorTotal) if consumo.c_valorTotal else '$0.00',
                estado,
                dias_transcurridos,
                truncar_texto(consumo.c_observaciones or 'Sin observaciones', 30),
                clasificacion
            ])
        
        # Crear tabla detallada de asignaciones
        current_row = crear_tabla_detallada_excel(ws, asig_headers, asig_data, current_row, "HISTORIAL DETALLADO DE ASIGNACIONES")
        
        # Análisis estadístico de asignaciones
        current_row += 2
        ws[f'A{current_row}'] = "ANÁLISIS DE ASIGNACIONES DEL PERSONAL"
        ws[f'A{current_row}'].font = Font(name='Calibri', size=14, bold=True, color="2E5984")
        ws.merge_cells(f'A{current_row}:F{current_row}')
        current_row += 2
        
        analisis_headers = ['Métrica', 'Valor', 'Porcentaje', 'Observaciones']
        analisis_data = [
            ['Total de Asignaciones', len(consumos), '100%', 'Asignaciones registradas'],
            ['Cantidad Total Asignada', cantidad_total, 'N/A', 'Unidades totales asignadas'],
            ['Valor Total Asignado', formatear_valor_moneda(valor_total_asignaciones), '100%', 'Valor monetario total'],
            ['Promedio por Asignación', formatear_valor_moneda(valor_total_asignaciones/len(consumos)) if consumos else '$0.00', 'N/A', 'Valor promedio por asignación']
        ]
        
        # Agregar desglose por estado
        for estado, cantidad in asignaciones_por_estado.items():
            porcentaje = f"{(cantidad/len(consumos)*100):.1f}%" if consumos else "0%"
            analisis_data.append([f'Asignaciones "{estado}"', cantidad, porcentaje, f'Estado: {estado}'])
        
        current_row = crear_tabla_detallada_excel(ws, analisis_headers, analisis_data, current_row)
    
    # Ajustar ancho de columnas
    ajustar_columnas_excel(ws)
    
    # Crear respuesta
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    response = make_response(output.read())
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Content-Disposition'] = f'attachment; filename=CNM_Detalle_{persona.pe_nombre}_Detallado_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    
    return response

def _exportar_detalle_personal_pdf(persona, consumos):
    """Exporta detalle de personal a PDF con formato detallado y análisis estadístico"""
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4,
                          rightMargin=0.5*inch, leftMargin=0.5*inch,
                          topMargin=0.5*inch, bottomMargin=0.5*inch)
    
    elements = []
    
    # Crear cabecera institucional
    elements.extend(crear_cabecera_pdf(f"Detalle de Personal - {persona.pe_nombre}"))
    
    # Información personal detallada
    info_headers = ['Campo', 'Valor', 'Estado']
    info_data = [
        ['Código', persona.pe_codigo or 'N/A', 'Asignado' if persona.pe_codigo else 'Sin asignar'],
        ['Nombre Completo', f"{persona.pe_nombre or ''} {persona.pe_apellido or ''}".strip(), 'Completo' if persona.pe_nombre else 'Incompleto'],
        ['Cédula de Identidad', persona.pe_ci or 'N/A', 'Registrada' if persona.pe_ci else 'Faltante'],
        ['Teléfono', persona.pe_telefono or 'N/A', 'Registrado' if persona.pe_telefono else 'Faltante'],
        ['Correo Electrónico', truncar_texto(persona.pe_correo or 'N/A', 25), 'Registrado' if persona.pe_correo else 'Faltante'],
        ['Dirección', truncar_texto(persona.pe_direccion or 'N/A', 25), 'Registrada' if persona.pe_direccion else 'Faltante'],
        ['Cargo', persona.pe_cargo or 'N/A', 'Asignado' if persona.pe_cargo else 'Sin asignar'],
        ['Estado', persona.pe_estado or 'N/A', persona.pe_estado or 'Sin definir']
    ]
    
    elements.extend(crear_tabla_detallada_pdf(
        info_headers, info_data,
        "INFORMACIÓN PERSONAL DETALLADA",
        [1.5*inch, 2.5*inch, 1.5*inch]
    ))
    
    # Análisis de asignaciones si existen
    if consumos:
        # Calcular estadísticas
        valor_total_asignaciones = sum(float(consumo.c_valorTotal) if consumo.c_valorTotal else 0 for consumo in consumos)
        cantidad_total = sum(consumo.c_cantidad or 0 for consumo in consumos)
        asignaciones_por_estado = {}
        
        for consumo in consumos:
            estado = consumo.c_estado or 'Sin estado'
            asignaciones_por_estado[estado] = asignaciones_por_estado.get(estado, 0) + 1
        
        # Resumen de asignaciones
        resumen_headers = ['Métrica', 'Valor', 'Observaciones']
        resumen_data = [
            ['Total Asignaciones', str(len(consumos)), 'Asignaciones registradas'],
            ['Cantidad Total', str(cantidad_total), 'Unidades asignadas'],
            ['Valor Total', formatear_valor_moneda(valor_total_asignaciones), 'Valor monetario total'],
            ['Promedio por Asignación', formatear_valor_moneda(valor_total_asignaciones/len(consumos)) if consumos else '$0.00', 'Valor promedio']
        ]
        
        # Agregar estados
        for estado, cantidad in asignaciones_por_estado.items():
            porcentaje = f"({(cantidad/len(consumos)*100):.1f}%)" if consumos else "(0%)"
            resumen_data.append([f'Estado "{estado}"', f"{cantidad} {porcentaje}", f'Asignaciones en {estado.lower()}'])
        
        elements.extend(crear_tabla_detallada_pdf(
            resumen_headers, resumen_data,
            "RESUMEN DE ASIGNACIONES",
            [1.5*inch, 1.5*inch, 2.5*inch]
        ))
        
        # Historial detallado de asignaciones
        asig_headers = ['Fecha', 'Artículo', 'Cant.', 'Valor', 'Estado', 'Días']
        asig_data = []
        
        for consumo in consumos:
            # Calcular días transcurridos
            if consumo.c_fecha:
                if isinstance(consumo.c_fecha, date):
                    fecha_asignacion = consumo.c_fecha
                else:
                    fecha_asignacion = consumo.c_fecha.date()
                dias_transcurridos = (datetime.now().date() - fecha_asignacion).days
            else:
                dias_transcurridos = 0
            
            asig_data.append([
                formatear_fecha(consumo.c_fecha) if consumo.c_fecha else 'N/A',
                truncar_texto(consumo.item.i_nombre if consumo.item else 'N/A', 15),
                str(consumo.c_cantidad or 0),
                formatear_valor_moneda(consumo.c_valorTotal) if consumo.c_valorTotal else '$0.00',
                consumo.c_estado or 'N/A',
                str(dias_transcurridos)
            ])
        
        elements.extend(crear_tabla_detallada_pdf(
            asig_headers, asig_data,
            "HISTORIAL DETALLADO DE ASIGNACIONES",
            [0.8*inch, 1.5*inch, 0.5*inch, 0.8*inch, 0.8*inch, 0.5*inch]
        ))
    
    # Construir PDF
    doc.build(elements)
    buffer.seek(0)
    
    response = make_response(buffer.read())
    response.headers['Content-Type'] = 'application/pdf'
    response.headers['Content-Disposition'] = f'attachment; filename=CNM_Detalle_{persona.pe_nombre}_Detallado_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf'
    
    return response