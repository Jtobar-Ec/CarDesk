from flask import Blueprint, render_template, request, redirect, url_for, flash, make_response
from flask_login import login_required, current_user
from app.services import ArticuloService
from app.services.proveedor_service import ProveedorService
from sqlalchemy import desc, asc
from app.database.models import Articulo, Item, Consumo, Persona
from app.database import db
import io
from openpyxl import Workbook
from reportlab.lib.pagesizes import A4, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from reportlab.lib.units import inch
from datetime import datetime
from app.utils.export_utils import (
    crear_cabecera_excel, crear_estilos_excel, crear_cabecera_pdf,
    crear_estilos_pdf, aplicar_estilo_tabla_pdf, ajustar_columnas_excel,
    crear_tabla_detallada_excel, crear_tabla_detallada_pdf,
    formatear_valor_moneda, formatear_fecha, truncar_texto
)
from openpyxl.styles import Font

bp = Blueprint('articulos', __name__)
articulo_service = ArticuloService()
proveedor_service = ProveedorService()

@bp.route('/')
@login_required
def listar_articulos():
    """Lista todos los artículos con filtros, paginación y exportación"""
    # Parámetros de filtrado
    q = request.args.get('q', '').strip()
    estado_stock = request.args.get('estado_stock', '')
    orden = request.args.get('orden', 'reciente')
    per_page = int(request.args.get('per_page', 25))
    page = int(request.args.get('page', 1))
    export_format = request.args.get('export')
    
    # Query base
    query = db.session.query(Articulo, Item).join(Item, Articulo.i_id == Item.id)
    
    # Aplicar filtros
    if q:
        query = query.filter(
            db.or_(
                Item.i_nombre.ilike(f'%{q}%'),
                Item.i_codigo.ilike(f'%{q}%')
            )
        )
    
    if estado_stock:
        if estado_stock == 'critico':
            query = query.filter(Item.i_cantidad < Articulo.a_stockMin)
        elif estado_stock == 'bajo':
            query = query.filter(
                db.and_(
                    Item.i_cantidad >= Articulo.a_stockMin,
                    Item.i_cantidad <= (Articulo.a_stockMin * 1.2)
                )
            )
        elif estado_stock == 'normal':
            query = query.filter(Item.i_cantidad > (Articulo.a_stockMin * 1.2))
    
    # Aplicar ordenamiento
    if orden == 'nombre':
        query = query.order_by(asc(Item.i_nombre))
    elif orden == 'stock':
        query = query.order_by(desc(Item.i_cantidad))
    elif orden == 'valor':
        query = query.order_by(desc(Item.i_vTotal))
    else:  # reciente
        query = query.order_by(desc(Item.created_at))
    
    # Manejar exportaciones
    if export_format in ['excel', 'pdf']:
        articulos = query.all()
        if export_format == 'excel':
            return _exportar_excel(articulos)
        else:
            return _exportar_pdf(articulos)
    
    # Paginación
    pagination = query.paginate(
        page=page, per_page=per_page, error_out=False
    )
    
    # Obtener proveedores para los modales de entrada
    proveedores = proveedor_service.obtener_todos()
    
    return render_template('articulos/list.html',
                         articulos=pagination.items,
                         pagination=pagination,
                         proveedores=proveedores)

def _exportar_excel(articulos):
    """Exporta artículos a Excel con formato detallado y profesional"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Artículos Detallado"
    
    # Configurar orientación horizontal
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    
    # Crear cabecera institucional
    current_row = crear_cabecera_excel(ws, "Reporte Detallado de Artículos")
    
    # Preparar datos detallados con estadísticas
    headers = [
        'Código', 'Nombre del Artículo', 'Stock Actual', 'Valor Unitario', 'Valor Total',
        'Stock Mínimo', 'Stock Máximo', 'Estado de Stock', 'Cuenta Contable',
        'Fecha Creación', 'Diferencia vs Stock Min', 'Rotación Sugerida'
    ]
    
    data = []
    total_valor = 0
    articulos_criticos = 0
    articulos_bajos = 0
    
    for articulo, item in articulos:
        # Determinar estado detallado
        if item.i_cantidad < articulo.a_stockMin:
            estado = "CRÍTICO"
            articulos_criticos += 1
        elif item.i_cantidad <= (articulo.a_stockMin * 1.2):
            estado = "BAJO"
            articulos_bajos += 1
        else:
            estado = "NORMAL"
        
        # Calcular diferencia con stock mínimo
        diferencia = item.i_cantidad - articulo.a_stockMin
        
        # Sugerir rotación
        if diferencia < 0:
            rotacion = "URGENTE - Reabastecer"
        elif diferencia <= 5:
            rotacion = "Próximo a reabastecer"
        else:
            rotacion = "Stock suficiente"
        
        total_valor += float(item.i_vTotal)
        
        data.append([
            item.i_codigo,
            item.i_nombre,
            item.i_cantidad,
            formatear_valor_moneda(item.i_vUnitario),
            formatear_valor_moneda(item.i_vTotal),
            articulo.a_stockMin,
            articulo.a_stockMax,
            estado,
            articulo.a_c_contable or 'N/A',
            formatear_fecha(item.created_at),
            diferencia,
            rotacion
        ])
    
    # Crear tabla detallada
    current_row = crear_tabla_detallada_excel(ws, headers, data, current_row, "INVENTARIO DETALLADO DE ARTÍCULOS")
    
    # Agregar resumen estadístico completo
    current_row += 2
    ws[f'A{current_row}'] = "ANÁLISIS ESTADÍSTICO DEL INVENTARIO"
    ws[f'A{current_row}'].font = Font(name='Calibri', size=14, bold=True, color="2E5984")
    ws.merge_cells(f'A{current_row}:F{current_row}')
    current_row += 2
    
    resumen_headers = ['Métrica', 'Valor', 'Porcentaje', 'Observaciones']
    resumen_data = [
        ['Total de Artículos', len(articulos), '100%', 'Inventario completo'],
        ['Artículos en Estado Crítico', articulos_criticos, f"{(articulos_criticos/len(articulos)*100):.1f}%" if articulos else "0%", 'Requieren atención inmediata'],
        ['Artículos con Stock Bajo', articulos_bajos, f"{(articulos_bajos/len(articulos)*100):.1f}%" if articulos else "0%", 'Próximos a reabastecimiento'],
        ['Artículos con Stock Normal', len(articulos)-articulos_criticos-articulos_bajos, f"{((len(articulos)-articulos_criticos-articulos_bajos)/len(articulos)*100):.1f}%" if articulos else "0%", 'Stock adecuado'],
        ['Valor Total del Inventario', formatear_valor_moneda(total_valor), '100%', 'Valor total en stock'],
        ['Promedio Valor por Artículo', formatear_valor_moneda(total_valor/len(articulos)) if articulos else "$0.00", 'N/A', 'Valor promedio unitario']
    ]
    
    current_row = crear_tabla_detallada_excel(ws, resumen_headers, resumen_data, current_row)
    
    # Ajustar ancho de columnas
    ajustar_columnas_excel(ws)
    
    # Crear respuesta
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    response = make_response(output.read())
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Content-Disposition'] = f'attachment; filename=CNM_Articulos_Detallado_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    
    return response

def _exportar_pdf(articulos):
    """Exporta artículos a PDF con formato detallado y profesional"""
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4),
                          rightMargin=0.5*inch, leftMargin=0.5*inch,
                          topMargin=0.5*inch, bottomMargin=0.5*inch)
    
    elements = []
    
    # Crear cabecera institucional
    elements.extend(crear_cabecera_pdf("Reporte Detallado de Artículos"))
    
    # Calcular estadísticas
    total_valor = sum(float(item.i_vTotal) for _, item in articulos)
    articulos_criticos = sum(1 for articulo, item in articulos if item.i_cantidad < articulo.a_stockMin)
    articulos_bajos = sum(1 for articulo, item in articulos if articulo.a_stockMin <= item.i_cantidad <= (articulo.a_stockMin * 1.2))
    
    # Resumen estadístico
    resumen_headers = ['Métrica', 'Cantidad', 'Porcentaje', 'Valor']
    resumen_data = [
        ['Total Artículos', str(len(articulos)), '100%', formatear_valor_moneda(total_valor)],
        ['Estado Crítico', str(articulos_criticos), f"{(articulos_criticos/len(articulos)*100):.1f}%" if articulos else "0%", 'Requiere atención'],
        ['Stock Bajo', str(articulos_bajos), f"{(articulos_bajos/len(articulos)*100):.1f}%" if articulos else "0%", 'Próximo reabastecimiento'],
        ['Stock Normal', str(len(articulos)-articulos_criticos-articulos_bajos), f"{((len(articulos)-articulos_criticos-articulos_bajos)/len(articulos)*100):.1f}%" if articulos else "0%", 'Stock adecuado']
    ]
    
    elements.extend(crear_tabla_detallada_pdf(
        resumen_headers, resumen_data,
        "RESUMEN ESTADÍSTICO DEL INVENTARIO",
        [2*inch, 1*inch, 1*inch, 1.5*inch]
    ))
    
    # Datos detallados de la tabla
    data_headers = ['Código', 'Nombre', 'Stock', 'V.Unit', 'V.Total', 'Min/Max', 'Estado']
    data = []
    
    for articulo, item in articulos:
        # Determinar estado
        if item.i_cantidad < articulo.a_stockMin:
            estado = "CRÍTICO"
        elif item.i_cantidad <= (articulo.a_stockMin * 1.2):
            estado = "BAJO"
        else:
            estado = "NORMAL"
        
        data.append([
            item.i_codigo,
            truncar_texto(item.i_nombre, 20),
            str(item.i_cantidad),
            formatear_valor_moneda(item.i_vUnitario),
            formatear_valor_moneda(item.i_vTotal),
            f"{articulo.a_stockMin}/{articulo.a_stockMax}",
            estado
        ])
    
    # Crear tabla detallada de artículos
    elements.extend(crear_tabla_detallada_pdf(
        data_headers, data,
        "INVENTARIO DETALLADO DE ARTÍCULOS",
        [0.8*inch, 2*inch, 0.6*inch, 0.8*inch, 0.8*inch, 0.8*inch, 0.8*inch]
    ))
    
    # Construir PDF
    doc.build(elements)
    buffer.seek(0)
    
    response = make_response(buffer.read())
    response.headers['Content-Type'] = 'application/pdf'
    response.headers['Content-Disposition'] = f'attachment; filename=CNM_Articulos_Detallado_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf'
    
    return response
    
    for articulo, item in articulos:
        # Determinar estado
        if item.i_cantidad < articulo.a_stockMin:
            estado = "Crítico"
        elif item.i_cantidad <= (articulo.a_stockMin * 1.2):
            estado = "Bajo"
        else:
            estado = "Normal"
        
        data.append([
            item.i_codigo,
            item.i_nombre[:25] + '...' if len(item.i_nombre) > 25 else item.i_nombre,
            str(item.i_cantidad),
            f"${float(item.i_vUnitario):.2f}",
            f"${float(item.i_vTotal):.2f}",
            f"{articulo.a_stockMin}/{articulo.a_stockMax}",
            estado,
            articulo.a_c_contable
        ])
    
    # Crear tabla
    table = Table(data, colWidths=[0.8*inch, 2.2*inch, 0.6*inch, 0.8*inch, 0.8*inch, 0.8*inch, 0.7*inch, 0.8*inch])
    
    # Aplicar estilo estandarizado
    aplicar_estilo_tabla_pdf(table)
    
    elements.append(table)
    
    # Construir PDF
    doc.build(elements)
    buffer.seek(0)
    
    response = make_response(buffer.read())
    response.headers['Content-Type'] = 'application/pdf'
    response.headers['Content-Disposition'] = f'attachment; filename=CNM_Articulos_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf'
    
    return response

@bp.route('/nuevo', methods=['GET', 'POST'])
@login_required
def nuevo_articulo():
    """Crear múltiples artículos por proveedor en una factura"""
    if request.method == 'POST':
        try:
            # Validar datos según si es donación o compra
            es_donacion = request.form.get('es_donacion') == 'on'
            observaciones_donacion = request.form.get('observaciones_donacion', '').strip()
            
            if not es_donacion:
                # Solo validar proveedor y factura si NO es donación
                proveedor_id = request.form.get('proveedor_id')
                numero_factura = request.form.get('numero_factura', '').strip()
                
                if not proveedor_id:
                    flash('Debe seleccionar un proveedor para compras', 'error')
                    return redirect(url_for('articulos.nuevo_articulo'))
                
                if not numero_factura:
                    flash('Debe ingresar el número de factura para compras', 'error')
                    return redirect(url_for('articulos.nuevo_articulo'))
                    
                proveedor_id = int(proveedor_id)
            else:
                # Para donaciones no se requiere proveedor ni factura
                proveedor_id = None
                numero_factura = "DONACIÓN"
            
            if not es_donacion:
                proveedor_id = int(proveedor_id)
                
                # Verificar que el proveedor existe
                proveedor = proveedor_service.obtener_por_id(proveedor_id)
                if not proveedor:
                    flash('El proveedor seleccionado no existe', 'error')
                    return redirect(url_for('articulos.nuevo_articulo'))
            else:
                proveedor = None
            
            articulos_data = []
            articulos_procesados = set()
            
            # Procesar múltiples artículos del formulario
            i = 0
            while f'nombre_{i}' in request.form:
                nombre = request.form.get(f'nombre_{i}', '').strip()
                cuenta_contable = request.form.get(f'cuenta_contable_{i}', '').strip()
                cantidad = request.form.get(f'cantidad_{i}')
                valor_unitario = request.form.get(f'valor_unitario_{i}')
                stock_min = request.form.get(f'stock_min_{i}', '5')
                stock_max = request.form.get(f'stock_max_{i}', '100')
                
                if nombre and cuenta_contable and cantidad and valor_unitario:
                    try:
                        cantidad = int(cantidad)
                        valor_unitario = float(valor_unitario)
                        stock_min = int(stock_min) if stock_min else 5
                        stock_max = int(stock_max) if stock_max else 100
                        
                        # Validaciones básicas
                        if cantidad <= 0:
                            flash(f'La cantidad del artículo "{nombre}" debe ser mayor a 0', 'error')
                            return redirect(url_for('articulos.nuevo_articulo'))
                        
                        if valor_unitario <= 0:
                            flash(f'El valor unitario del artículo "{nombre}" debe ser mayor a 0', 'error')
                            return redirect(url_for('articulos.nuevo_articulo'))
                        
                        if stock_min < 0 or stock_max <= 0 or stock_max <= stock_min:
                            flash(f'Los valores de stock del artículo "{nombre}" no son válidos', 'error')
                            return redirect(url_for('articulos.nuevo_articulo'))
                        
                        # Verificar duplicados por nombre
                        nombre_lower = nombre.lower()
                        if nombre_lower in articulos_procesados:
                            flash(f'No puede agregar el mismo artículo "{nombre}" múltiples veces', 'error')
                            return redirect(url_for('articulos.nuevo_articulo'))
                        
                        # Obtener campos opcionales
                        serial = request.form.get(f'serial_{i}', '').strip()
                        codigo_identificacion = request.form.get(f'codigo_identificacion_{i}', '').strip()
                        
                        articulos_data.append({
                            'nombre': nombre,
                            'cuenta_contable': cuenta_contable,
                            'cantidad': cantidad,
                            'valor_unitario': valor_unitario,
                            'stock_min': stock_min,
                            'stock_max': stock_max,
                            'serial': serial if serial else None,
                            'codigo_identificacion': codigo_identificacion if codigo_identificacion else None
                        })
                        articulos_procesados.add(nombre_lower)
                        
                    except (ValueError, TypeError) as e:
                        flash(f'Error en los datos del artículo "{nombre}": valores inválidos', 'error')
                        return redirect(url_for('articulos.nuevo_articulo'))
                
                i += 1
            
            # Validar que se hayan agregado artículos
            if not articulos_data:
                flash('Debe agregar al menos un artículo válido', 'error')
                return redirect(url_for('articulos.nuevo_articulo'))
            
            # Procesar todos los artículos
            articulos_creados = 0
            errores = []
            
            try:
                # Usar transacción para asegurar consistencia
                for articulo_data in articulos_data:
                    try:
                        # Crear el artículo (ya incluye el stock inicial)
                        articulo, item = articulo_service.crear_articulo(
                            nombre=articulo_data['nombre'],
                            cantidad=articulo_data['cantidad'],
                            valor_unitario=articulo_data['valor_unitario'],
                            cuenta_contable=articulo_data['cuenta_contable'],
                            stock_min=articulo_data['stock_min'],
                            stock_max=articulo_data['stock_max'],
                            usuario_id=current_user.id,
                            serial=articulo_data.get('serial'),
                            codigo_identificacion=articulo_data.get('codigo_identificacion')
                        )
                        
                        # CORRECCIÓN: No registrar entrada adicional
                        # El artículo ya se crea con el stock inicial correcto
                        # Solo registrar el movimiento de entrada inicial para auditoría
                        from app.database.models import MovimientoDetalle
                        from datetime import datetime
                        
                        # Preparar observaciones con información de donación
                        if es_donacion:
                            observaciones_entrada = f"Stock inicial - DONACIÓN - {articulo_data['cuenta_contable']}"
                            if observaciones_donacion:
                                observaciones_entrada += f" - {observaciones_donacion}"
                        else:
                            observaciones_entrada = f"Stock inicial - Factura: {numero_factura} - {articulo_data['cuenta_contable']}"
                        
                        # Registrar movimiento de entrada inicial para auditoría (sin duplicar stock)
                        movimiento = MovimientoDetalle(
                            m_fecha=datetime.now().date(),
                            m_tipo='entrada',
                            m_cantidad=articulo_data['cantidad'],
                            m_valorUnitario=articulo_data['valor_unitario'],
                            m_valorTotal=articulo_data['cantidad'] * articulo_data['valor_unitario'],
                            m_observaciones=observaciones_entrada,
                            i_id=item.id,
                            u_id=current_user.id
                        )
                        db.session.add(movimiento)
                        
                        articulos_creados += 1
                        
                    except Exception as e:
                        errores.append(f"{articulo_data['nombre']}: {str(e)}")
                        db.session.rollback()
                        continue
                
                # Mostrar resultados
                if articulos_creados > 0:
                    if es_donacion:
                        flash(f'Se registraron {articulos_creados} artículo(s) como DONACIÓN exitosamente', 'success')
                    else:
                        flash(f'Se crearon {articulos_creados} artículo(s) exitosamente - COMPRA del proveedor {proveedor.p_razonsocial} - Factura: {numero_factura}', 'success')
                
                if errores:
                    for error in errores:
                        flash(f'Error: {error}', 'warning')
                
                # Redirigir según el resultado
                if articulos_creados > 0:
                    return redirect(url_for('articulos.listar_articulos'))
                else:
                    return redirect(url_for('articulos.nuevo_articulo'))
                    
            except Exception as e:
                db.session.rollback()
                flash(f'Error general en el procesamiento: {str(e)}', 'error')
                return redirect(url_for('articulos.nuevo_articulo'))
            
        except Exception as e:
            flash(f'Error al crear artículos: {str(e)}', 'error')
            return redirect(url_for('articulos.nuevo_articulo'))
    
    # GET: Mostrar formulario
    try:
        # Obtener todos los proveedores para el formulario
        proveedores = proveedor_service.obtener_todos()
        # Obtener artículos existentes para autocompletado
        articulos_existentes = articulo_service.obtener_todos()
        return render_template('articulos/form.html',
                             proveedores=proveedores,
                             articulos_existentes=articulos_existentes)
        
    except Exception as e:
        flash(f'Error al cargar el formulario: {str(e)}', 'error')
        return redirect(url_for('articulos.listar_articulos'))
@bp.route('/<int:articulo_id>/detalle')
@login_required
def detalle_articulo(articulo_id):
    """Ver detalles de un artículo"""
    export_format = request.args.get('export')
    
    resultado = articulo_service.obtener_articulo_por_id(articulo_id)
    if not resultado:
        flash('Artículo no encontrado', 'error')
        return redirect(url_for('articulos.listar_articulos'))
    
    articulo = resultado
    item = articulo.item
    
    # Parámetros de filtrado para movimientos
    tipo_mov = request.args.get('tipo_mov', '')
    fecha_desde_mov = request.args.get('fecha_desde_mov', '')
    fecha_hasta_mov = request.args.get('fecha_hasta_mov', '')
    page_mov = int(request.args.get('page_mov', 1))
    per_page_mov = int(request.args.get('per_page_mov', 10))
    
    # Obtener TODOS los movimientos para cálculos (sin filtros)
    from app.database.models import MovimientoDetalle
    todos_movimientos = db.session.query(MovimientoDetalle).filter_by(
        i_id=item.id
    ).order_by(MovimientoDetalle.m_fecha.asc(), MovimientoDetalle.id.asc()).all()
    
    # Calcular valores totales con TODOS los movimientos
    stock_actual = 0
    saldo_calculado = 0
    valor_total_historico = 0
    valor_total_actual = 0
    
    for mov in todos_movimientos:
        mov.m_stock_anterior = stock_actual
        
        if mov.m_tipo == 'entrada':
            stock_actual += mov.m_cantidad
            saldo_calculado += mov.m_cantidad
            valor_total_historico += mov.m_valorTotal
            valor_total_actual += mov.m_valorTotal
        elif mov.m_tipo == 'salida':
            stock_actual -= mov.m_cantidad
            valor_total_actual -= mov.m_valorTotal
        
        mov.m_stock_actual = stock_actual
    
    # Query para movimientos con filtros y paginación
    query_mov = db.session.query(MovimientoDetalle).filter_by(i_id=item.id)
    
    if tipo_mov:
        query_mov = query_mov.filter(MovimientoDetalle.m_tipo == tipo_mov)
    
    if fecha_desde_mov:
        try:
            from datetime import datetime
            fecha_desde_obj = datetime.strptime(fecha_desde_mov, '%Y-%m-%d').date()
            query_mov = query_mov.filter(MovimientoDetalle.m_fecha >= fecha_desde_obj)
        except ValueError:
            pass
    
    if fecha_hasta_mov:
        try:
            from datetime import datetime
            fecha_hasta_obj = datetime.strptime(fecha_hasta_mov, '%Y-%m-%d').date()
            query_mov = query_mov.filter(MovimientoDetalle.m_fecha <= fecha_hasta_obj)
        except ValueError:
            pass
    
    # Paginación de movimientos
    pagination_mov = query_mov.order_by(MovimientoDetalle.m_fecha.desc()).paginate(
        page=page_mov, per_page=per_page_mov, error_out=False
    )
    
    # Aplicar cálculos de stock a movimientos paginados
    movimientos_dict = {mov.id: mov for mov in todos_movimientos}
    for mov in pagination_mov.items:
        if mov.id in movimientos_dict:
            mov.m_stock_anterior = movimientos_dict[mov.id].m_stock_anterior
            mov.m_stock_actual = movimientos_dict[mov.id].m_stock_actual
    
    # Obtener auditoría de cambios (movimientos que registran cambios)
    auditoria = db.session.query(MovimientoDetalle).filter_by(
        i_id=item.id
    ).filter(
        MovimientoDetalle.m_tipo.in_(['ajuste_precio', 'entrada', 'salida'])
    ).order_by(MovimientoDetalle.m_fecha.desc()).limit(20).all()
    
    # Manejar exportaciones
    if export_format in ['excel', 'pdf']:
        if export_format == 'excel':
            return _exportar_articulos_excel(articulo, item, todos_movimientos, saldo_calculado)
        else:
            return _exportar_articulos_pdf(articulo, item, todos_movimientos, saldo_calculado)
    
    # Obtener todos los proveedores para el modal
    proveedores = proveedor_service.obtener_todos()
    
    return render_template('articulos/detail.html',
                         articulo=articulo, item=item, movimientos=pagination_mov.items,
                         pagination_mov=pagination_mov, saldo_calculado=saldo_calculado,
                         auditoria=auditoria, proveedores=proveedores,
                         valor_total_historico=valor_total_historico,
                         valor_total_actual=valor_total_actual)

@bp.route('/buscar')
def buscar_articulos():
    """Busca artículos por nombre o código"""
    termino = request.args.get('q', '').strip()
    if termino:
        articulos = articulo_service.buscar_articulos(termino)
    else:
        articulos = articulo_service.obtener_todos_articulos()
    
    return render_template('articulos/list.html', articulos=articulos, termino_busqueda=termino)

@bp.route('/<int:articulo_id>/editar', methods=['GET', 'POST'])
def editar_articulo(articulo_id):
    """Editar un artículo"""
    resultado = articulo_service.obtener_articulo_por_id(articulo_id)
    if not resultado:
        flash('Artículo no encontrado', 'error')
        return redirect(url_for('articulos.listar_articulos'))
    
    # Obtener tanto el artículo como el item
    articulo = resultado
    item = articulo.item
    
    if request.method == 'POST':
        try:
            articulo_service.actualizar_articulo(
                articulo_id,
                nombre=request.form['nombre'],
                codigo=request.form['codigo'],
                cuenta_contable=request.form['cuenta_contable'],
                stock_min=int(request.form.get('stock_min', 0)),
                stock_max=int(request.form.get('stock_max', 100))
            )
            
            flash('Artículo actualizado exitosamente', 'success')
            return redirect(url_for('articulos.detalle_articulo', articulo_id=articulo_id))
        except Exception as e:
            flash(f'Error al actualizar artículo: {str(e)}', 'error')
    
    # Obtener todos los proveedores para el formulario (aunque no se use en edición)
    proveedores = proveedor_service.obtener_todos()
    return render_template('articulos/form.html', articulo=articulo, item=item, proveedores=proveedores)

@bp.route('/<int:articulo_id>/eliminar', methods=['POST'])
def eliminar_articulo(articulo_id):
    """Eliminar un artículo"""
    try:
        if articulo_service.eliminar_articulo(articulo_id):
            flash('Artículo eliminado exitosamente', 'success')
        else:
            flash('No se pudo eliminar el artículo', 'error')
    except Exception as e:
        flash(f'Error al eliminar artículo: {str(e)}', 'error')
    
    return redirect(url_for('articulos.listar_articulos'))

@bp.route('/<int:articulo_id>/actualizar-precio', methods=['POST'])
def actualizar_precio(articulo_id):
    """Actualiza el valor unitario de un artículo"""
    try:
        nuevo_valor = float(request.form['nuevo_valor'])
        observaciones = request.form.get('observaciones', '').strip()
        usuario_id = 1  # Por ahora usuario fijo
        
        articulo_service.actualizar_valor_unitario(
            articulo_id, nuevo_valor, usuario_id, observaciones
        )
        
        flash('Valor unitario actualizado exitosamente', 'success')
    except Exception as e:
        flash(f'Error al actualizar precio: {str(e)}', 'error')
    
    return redirect(request.referrer or url_for('articulos.listar_articulos'))

@bp.route('/stock-bajo')
@login_required
def stock_bajo():
    """Lista artículos con stock bajo"""
    articulos = articulo_service.obtener_articulos_stock_bajo()
    return render_template('articulos/stock_bajo.html', articulos=articulos)

@bp.route('/<int:articulo_id>/entrada', methods=['POST'])
@login_required
def registrar_entrada(articulo_id):
    """Registra una entrada de artículo"""
    try:
        cantidad = int(request.form['cantidad'])
        valor_unitario = float(request.form['valor_unitario'])
        proveedor_id = request.form.get('proveedor_id')
        numero_factura = request.form.get('numero_factura', '').strip()
        observaciones = request.form.get('observaciones')
        usuario_id = current_user.id
        
        # Convertir proveedor_id a int si existe
        if proveedor_id and proveedor_id.strip():
            proveedor_id = int(proveedor_id)
        else:
            proveedor_id = None
        
        # Preparar observaciones con número de factura
        observaciones_final = observaciones or ""
        if numero_factura:
            observaciones_final = f"Factura: {numero_factura}" + (f" - {observaciones}" if observaciones else "")
        
        # Obtener el item_id correcto
        resultado = articulo_service.obtener_por_id(articulo_id)
        if not resultado:
            flash('Artículo no encontrado', 'error')
            return redirect(request.referrer or url_for('articulos.listar_articulos'))
        
        articulo, item = resultado
        
        from datetime import datetime
        articulo_service.registrar_entrada(
            item.id, cantidad, valor_unitario, usuario_id,
            proveedor_id=proveedor_id, observaciones=observaciones_final,
            fecha_hora=datetime.now()
        )
        
        flash('Entrada registrada exitosamente', 'success')
    except Exception as e:
        flash(f'Error al registrar entrada: {str(e)}', 'error')
    
    return redirect(request.referrer or url_for('articulos.listar_articulos'))

@bp.route('/<int:articulo_id>/cambiar-estado', methods=['POST'])
@login_required
def cambiar_estado(articulo_id):
    """Cambiar estado de un artículo (Activo, Dañado, Baja)"""
    try:
        # Obtener el artículo
        resultado = articulo_service.obtener_por_id(articulo_id)
        if not resultado:
            flash('Artículo no encontrado', 'error')
            return redirect(url_for('articulos.listar_articulos'))
        
        articulo, item = resultado
        
        nuevo_estado = request.form.get('nuevo_estado')
        observaciones_estado = request.form.get('observaciones_estado', '').strip()
        
        if not nuevo_estado or nuevo_estado not in ['Activo', 'Dañado', 'Baja']:
            flash('Estado inválido', 'error')
            return redirect(url_for('articulos.listar_articulos'))
        
        # Actualizar estado
        item.i_estado = nuevo_estado
        item.i_observaciones_estado = observaciones_estado if observaciones_estado else None
        
        db.session.commit()
        
        # Mensaje según el estado
        if nuevo_estado == 'Dañado':
            flash(f'Artículo "{item.i_nombre}" marcado como DAÑADO', 'warning')
        elif nuevo_estado == 'Baja':
            flash(f'Artículo "{item.i_nombre}" dado de BAJA', 'info')
        else:
            flash(f'Artículo "{item.i_nombre}" reactivado', 'success')
        
    except Exception as e:
        db.session.rollback()
        flash(f'Error al cambiar estado: {str(e)}', 'error')
    
    return redirect(url_for('articulos.listar_articulos'))

@bp.route('/asignaciones')
@login_required
def listar_asignaciones():
    """Lista todas las asignaciones de artículos a personal con filtros y paginación"""
    from app.database.models import Consumo, Persona, Item
    from app import db
    
    # Obtener filtros y paginación
    estado = request.args.get('estado', '').strip()
    articulo = request.args.get('articulo', '').strip()
    persona = request.args.get('persona', '').strip()
    page = request.args.get('page', 1, type=int)
    per_page = 20
    
    # Query base
    query = db.session.query(Consumo, Persona, Item).join(
        Persona, Consumo.pe_id == Persona.id
    ).join(
        Item, Consumo.i_id == Item.id
    )
    
    # Aplicar filtros
    if estado:
        query = query.filter(Consumo.c_estado == estado)
    
    if articulo and articulo.lower() != 'none':
        query = query.filter(
            db.or_(
                Item.i_nombre.ilike(f'%{articulo}%'),
                Item.i_codigo.ilike(f'%{articulo}%')
            )
        )
    
    if persona and persona.lower() != 'none':
        query = query.filter(
            db.or_(
                Persona.pe_nombre.ilike(f'%{persona}%'),
                Persona.pe_apellido.ilike(f'%{persona}%')
            )
        )
    
    # Ordenar por fecha descendente (más recientes primero) y paginar
    pagination = query.order_by(Consumo.c_fecha.desc(), Consumo.c_hora.desc())\
                     .paginate(page=page, per_page=per_page, error_out=False)
    
    return render_template('articulos/asignaciones.html',
                         asignaciones=pagination.items,
                         pagination=pagination)

@bp.route('/asignaciones/<int:persona_id>')
@login_required
def asignaciones_por_persona(persona_id):
    """Lista las asignaciones de una persona específica"""
    from app.database.models import Consumo, Persona, Item
    from app import db
    
    # Verificar que la persona existe
    persona = Persona.query.get(persona_id)
    if not persona:
        flash('Persona no encontrada', 'error')
        return redirect(url_for('articulos.listar_asignaciones'))
    
    # Paginación
    page = request.args.get('page', 1, type=int)
    per_page = 20
    
    # Obtener asignaciones de la persona con paginación
    pagination = db.session.query(Consumo, Persona, Item).join(
        Persona, Consumo.pe_id == Persona.id
    ).join(
        Item, Consumo.i_id == Item.id
    ).filter(Consumo.pe_id == persona_id).order_by(
        Consumo.c_fecha.desc(), Consumo.c_hora.desc()
    ).paginate(page=page, per_page=per_page, error_out=False)
    
    return render_template('articulos/asignaciones.html',
                         asignaciones=pagination.items,
                         pagination=pagination,
                         persona_filtro=persona)

@bp.route('/asignaciones/<int:consumo_id>/cambiar-estado', methods=['POST'])
@login_required
def cambiar_estado_asignacion(consumo_id):
    """Cambia el estado de una asignación con lógica de retorno de stock"""
    from app.database.models import Consumo, Item
    from datetime import datetime
    
    try:
        consumo = Consumo.query.get(consumo_id)
        if not consumo:
            flash('Asignación no encontrada', 'error')
            return redirect(url_for('articulos.listar_asignaciones'))
        
        # Verificar si puede editarse (48 horas)
        if not consumo.puede_editar_por_tiempo:
            flash('No se puede modificar esta asignación. Han pasado más de 48 horas desde su creación.', 'error')
            return redirect(request.referrer or url_for('articulos.listar_asignaciones'))
        
        nuevo_estado = request.form['nuevo_estado']
        observaciones = request.form.get('observaciones', '')
        estado_anterior = consumo.c_estado
        
        # Solo "Devuelto" retorna stock al inventario
        if nuevo_estado == 'Devuelto' and estado_anterior != 'Devuelto':
            item = Item.query.get(consumo.i_id)
            if item:
                # Solo retornar la cantidad física al stock
                item.i_cantidad += consumo.c_cantidad
                consumo.c_fecha_devolucion = datetime.utcnow()
        
        # Si se cambia de "Devuelto" a cualquier otro estado, descontar stock
        elif estado_anterior == 'Devuelto' and nuevo_estado != 'Devuelto':
            item = Item.query.get(consumo.i_id)
            if item:
                if item.i_cantidad >= consumo.c_cantidad:
                    item.i_cantidad -= consumo.c_cantidad
                    consumo.c_fecha_devolucion = None
                else:
                    flash(f'No hay suficiente stock disponible. Stock actual: {item.i_cantidad}', 'error')
                    return redirect(request.referrer or url_for('articulos.listar_asignaciones'))
        
        # Actualizar el estado
        consumo.c_estado = nuevo_estado
        
        # Agregar observaciones del cambio
        if observaciones:
            obs_anterior = consumo.c_observaciones or ''
            timestamp = datetime.now().strftime('%d/%m/%Y %H:%M')
            consumo.c_observaciones = f"{obs_anterior}\n[{timestamp} - {nuevo_estado}] {observaciones}".strip()
        
        db.session.commit()
        
        flash(f'Estado cambiado a "{nuevo_estado}" exitosamente', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error al cambiar estado: {str(e)}', 'error')
    
    return redirect(request.referrer or url_for('articulos.listar_asignaciones'))

@bp.route('/movimientos')
@login_required
def listar_movimientos():
    """Lista todos los movimientos de inventario con filtros y paginación"""
    from app.database.models import MovimientoDetalle, Item, Usuario, Entrada, Proveedor
    from datetime import datetime
    
    # Parámetros de filtrado
    buscar = request.args.get('buscar', '').strip()
    tipo = request.args.get('tipo', '')
    fecha_desde = request.args.get('fecha_desde', '')
    fecha_hasta = request.args.get('fecha_hasta', '')
    per_page = 20
    page = int(request.args.get('page', 1))
    export_format = request.args.get('export')
    
    # Query base con joins
    query = db.session.query(MovimientoDetalle, Item, Usuario).join(
        Item, MovimientoDetalle.i_id == Item.id
    ).join(
        Usuario, MovimientoDetalle.u_id == Usuario.id
    )
    
    # Aplicar filtros
    if buscar:
        query = query.filter(
            db.or_(
                Item.i_nombre.ilike(f'%{buscar}%'),
                Item.i_codigo.ilike(f'%{buscar}%'),
                MovimientoDetalle.m_observaciones.ilike(f'%{buscar}%')
            )
        )
    
    if tipo:
        query = query.filter(MovimientoDetalle.m_tipo == tipo)
    
    if fecha_desde:
        try:
            fecha_desde_obj = datetime.strptime(fecha_desde, '%Y-%m-%d').date()
            query = query.filter(MovimientoDetalle.m_fecha >= fecha_desde_obj)
        except ValueError:
            pass
    
    if fecha_hasta:
        try:
            fecha_hasta_obj = datetime.strptime(fecha_hasta, '%Y-%m-%d').date()
            query = query.filter(MovimientoDetalle.m_fecha <= fecha_hasta_obj)
        except ValueError:
            pass
    
    # Ordenar por fecha descendente
    query = query.order_by(desc(MovimientoDetalle.m_fecha), desc(MovimientoDetalle.id))
    
    # Manejar exportaciones
    if export_format in ['excel', 'pdf']:
        movimientos = query.all()
        if export_format == 'excel':
            return _exportar_movimientos_excel(movimientos)
        else:
            return _exportar_movimientos_pdf(movimientos)
    
    # Paginación
    pagination = query.paginate(
        page=page, per_page=per_page, error_out=False
    )
    
    return render_template('articulos/movimientos.html',
                         movimientos=pagination.items,
                         pagination=pagination)

def _exportar_movimientos_excel(movimientos):
    """Exporta movimientos a Excel con formato detallado y análisis estadístico"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Movimientos Detallado"
    
    # Configurar orientación horizontal
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    
    # Crear cabecera institucional
    current_row = crear_cabecera_excel(ws, "Reporte Detallado de Movimientos de Inventario")
    
    # Preparar datos detallados con estadísticas
    headers = [
        'Fecha', 'Hora', 'Artículo', 'Código', 'Tipo de Movimiento', 'Cantidad',
        'Valor Unitario', 'Valor Total', 'Usuario', 'Proveedor', 'Observaciones', 'Impacto'
    ]
    
    data = []
    total_entradas = 0
    total_salidas = 0
    valor_total_entradas = 0
    valor_total_salidas = 0
    movimientos_por_tipo = {}
    
    for movimiento, item, usuario in movimientos:
        # Obtener proveedor si existe
        proveedor_nombre = "N/A"
        if movimiento.e_id and movimiento.entrada and movimiento.entrada.proveedor:
            proveedor_nombre = movimiento.entrada.proveedor.p_razonsocial
        
        # Determinar impacto del movimiento
        if movimiento.m_tipo == 'entrada':
            impacto = "Incrementa stock"
            total_entradas += movimiento.m_cantidad
            valor_total_entradas += float(movimiento.m_valorTotal)
        elif movimiento.m_tipo == 'salida':
            impacto = "Reduce stock"
            total_salidas += movimiento.m_cantidad
            valor_total_salidas += float(movimiento.m_valorTotal)
        else:
            impacto = "Ajuste de inventario"
        
        # Contar movimientos por tipo
        tipo_mov = movimiento.m_tipo.title()
        movimientos_por_tipo[tipo_mov] = movimientos_por_tipo.get(tipo_mov, 0) + 1
        
        data.append([
            formatear_fecha(movimiento.m_fecha),
            movimiento.m_hora.strftime('%H:%M:%S') if hasattr(movimiento, 'm_hora') and movimiento.m_hora else 'N/A',
            item.i_nombre,
            item.i_codigo,
            tipo_mov,
            movimiento.m_cantidad,
            formatear_valor_moneda(movimiento.m_valorUnitario),
            formatear_valor_moneda(movimiento.m_valorTotal),
            usuario.u_username,
            truncar_texto(proveedor_nombre, 20),
            truncar_texto(movimiento.m_observaciones or 'Sin observaciones', 30),
            impacto
        ])
    
    # Crear tabla detallada
    current_row = crear_tabla_detallada_excel(ws, headers, data, current_row, "HISTORIAL DETALLADO DE MOVIMIENTOS")
    
    # Agregar análisis estadístico completo
    current_row += 2
    ws[f'A{current_row}'] = "ANÁLISIS ESTADÍSTICO DE MOVIMIENTOS"
    ws[f'A{current_row}'].font = Font(name='Calibri', size=14, bold=True, color="2E5984")
    ws.merge_cells(f'A{current_row}:F{current_row}')
    current_row += 2
    
    resumen_headers = ['Métrica', 'Cantidad', 'Valor', 'Porcentaje', 'Observaciones']
    resumen_data = [
        ['Total de Movimientos', len(movimientos), formatear_valor_moneda(valor_total_entradas + valor_total_salidas), '100%', 'Actividad total del inventario'],
        ['Total Entradas', total_entradas, formatear_valor_moneda(valor_total_entradas), f"{(total_entradas/(total_entradas+total_salidas)*100):.1f}%" if (total_entradas+total_salidas) > 0 else "0%", 'Incrementos de stock'],
        ['Total Salidas', total_salidas, formatear_valor_moneda(valor_total_salidas), f"{(total_salidas/(total_entradas+total_salidas)*100):.1f}%" if (total_entradas+total_salidas) > 0 else "0%", 'Reducciones de stock'],
        ['Diferencia Neta', total_entradas - total_salidas, formatear_valor_moneda(valor_total_entradas - valor_total_salidas), 'N/A', 'Balance de inventario'],
        ['Promedio por Movimiento', f"{(valor_total_entradas + valor_total_salidas)/len(movimientos):.2f}" if movimientos else "0", formatear_valor_moneda((valor_total_entradas + valor_total_salidas)/len(movimientos)) if movimientos else "$0.00", 'N/A', 'Valor promedio por transacción']
    ]
    
    # Agregar desglose por tipo de movimiento
    for tipo, cantidad in movimientos_por_tipo.items():
        porcentaje = f"{(cantidad/len(movimientos)*100):.1f}%" if movimientos else "0%"
        resumen_data.append([f'Movimientos tipo {tipo}', cantidad, 'N/A', porcentaje, f'Frecuencia de {tipo.lower()}'])
    
    current_row = crear_tabla_detallada_excel(ws, resumen_headers, resumen_data, current_row)
    
    # Ajustar ancho de columnas
    ajustar_columnas_excel(ws)
    
    # Crear respuesta
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    response = make_response(output.read())
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Content-Disposition'] = f'attachment; filename=CNM_Movimientos_Detallado_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    
    return response

def _exportar_movimientos_pdf(movimientos):
    """Exporta movimientos a PDF con formato detallado y análisis estadístico"""
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4),
                          rightMargin=0.5*inch, leftMargin=0.5*inch,
                          topMargin=0.5*inch, bottomMargin=0.5*inch)
    
    elements = []
    
    # Crear cabecera institucional
    elements.extend(crear_cabecera_pdf("Reporte Detallado de Movimientos de Inventario"))
    
    # Calcular estadísticas
    total_entradas = sum(mov.m_cantidad for mov, _, _ in movimientos if mov.m_tipo == 'entrada')
    total_salidas = sum(mov.m_cantidad for mov, _, _ in movimientos if mov.m_tipo == 'salida')
    valor_total_entradas = sum(float(mov.m_valorTotal) for mov, _, _ in movimientos if mov.m_tipo == 'entrada')
    valor_total_salidas = sum(float(mov.m_valorTotal) for mov, _, _ in movimientos if mov.m_tipo == 'salida')
    
    # Resumen estadístico
    resumen_headers = ['Métrica', 'Cantidad', 'Valor', 'Observaciones']
    resumen_data = [
        ['Total Movimientos', str(len(movimientos)), formatear_valor_moneda(valor_total_entradas + valor_total_salidas), 'Actividad total'],
        ['Total Entradas', str(total_entradas), formatear_valor_moneda(valor_total_entradas), 'Incrementos de stock'],
        ['Total Salidas', str(total_salidas), formatear_valor_moneda(valor_total_salidas), 'Reducciones de stock'],
        ['Balance Neto', str(total_entradas - total_salidas), formatear_valor_moneda(valor_total_entradas - valor_total_salidas), 'Diferencia entrada-salida']
    ]
    
    elements.extend(crear_tabla_detallada_pdf(
        resumen_headers, resumen_data,
        "RESUMEN ESTADÍSTICO DE MOVIMIENTOS",
        [1.5*inch, 1*inch, 1.2*inch, 2*inch]
    ))
    
    # Datos detallados de la tabla
    data_headers = ['Fecha', 'Artículo', 'Código', 'Tipo', 'Cant.', 'V.Unit', 'V.Total', 'Usuario']
    data = []
    
    for movimiento, item, usuario in movimientos:
        data.append([
            formatear_fecha(movimiento.m_fecha),
            truncar_texto(item.i_nombre, 18),
            item.i_codigo,
            movimiento.m_tipo.title(),
            str(movimiento.m_cantidad),
            formatear_valor_moneda(movimiento.m_valorUnitario),
            formatear_valor_moneda(movimiento.m_valorTotal),
            truncar_texto(usuario.u_username, 12)
        ])
    
    # Crear tabla detallada de movimientos
    elements.extend(crear_tabla_detallada_pdf(
        data_headers, data,
        "HISTORIAL DETALLADO DE MOVIMIENTOS",
        [0.8*inch, 1.8*inch, 0.8*inch, 0.7*inch, 0.5*inch, 0.7*inch, 0.8*inch, 0.8*inch]
    ))
    
    # Construir PDF
    doc.build(elements)
    buffer.seek(0)
    
    response = make_response(buffer.read())
    response.headers['Content-Type'] = 'application/pdf'
    response.headers['Content-Disposition'] = f'attachment; filename=CNM_Movimientos_Detallado_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf'
    
    return response

def _exportar_articulos_excel(articulo, item, movimientos, saldo_calculado):
    """Exporta detalle de artículo a Excel con cabecera institucional"""
    wb = Workbook()
    ws = wb.active
    ws.title = f"Detalle {item.i_codigo}"
    
    # Configurar orientación horizontal
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    
    # Crear cabecera institucional
    current_row = crear_cabecera_excel(ws, f"Detalle de Artículo - {item.i_codigo}")
    
    # Obtener estilos
    estilos = crear_estilos_excel()
    
    # Información del artículo
    info_headers = ['Campo', 'Valor']
    for col, header in enumerate(info_headers, 1):
        cell = ws.cell(row=current_row, column=col, value=header)
        cell.font = estilos['header_font']
        cell.fill = estilos['header_fill']
        cell.alignment = estilos['header_alignment']
        cell.border = estilos['header_border']
    current_row += 1
    
    # Datos del artículo
    info_data = [
        ['Código', item.i_codigo],
        ['Nombre', item.i_nombre],
        ['Stock Actual', item.i_cantidad],
        ['Saldo Calculado', saldo_calculado],
        ['Valor Unitario', f"${float(item.i_vUnitario):.2f}"],
        ['Stock Mínimo', articulo.a_stockMin],
        ['Stock Máximo', articulo.a_stockMax],
        ['Cuenta Contable', articulo.a_c_contable]
    ]
    
    for campo, valor in info_data:
        ws.cell(row=current_row, column=1, value=campo).font = estilos['data_font']
        ws.cell(row=current_row, column=2, value=valor).font = estilos['data_font']
        current_row += 1
    
    current_row += 2  # Espacio antes de movimientos
    
    # Título de movimientos
    ws.cell(row=current_row, column=1, value="HISTORIAL DE MOVIMIENTOS").font = estilos['header_font']
    current_row += 2
    
    # Encabezados de movimientos
    row_start = current_row
    headers = ['Fecha', 'Tipo', 'Cantidad', 'Valor Unit.', 'Valor Total', 'Usuario', 'Observaciones']
    
    # Escribir encabezados de movimientos
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row_start, column=col, value=header)
        cell.font = estilos['header_font']
        cell.fill = estilos['header_fill']
        cell.alignment = estilos['header_alignment']
        cell.border = estilos['header_border']
    row_start += 1
    
    # Escribir datos de movimientos
    for mov in movimientos:
        ws.cell(row=row_start, column=1, value=mov.m_fecha.strftime('%Y-%m-%d'))
        ws.cell(row=row_start, column=2, value=mov.m_tipo.title())
        ws.cell(row=row_start, column=3, value=mov.m_cantidad)
        ws.cell(row=row_start, column=4, value=float(mov.m_valorUnitario))
        ws.cell(row=row_start, column=5, value=float(mov.m_valorTotal))
        ws.cell(row=row_start, column=6, value=mov.u_id)
        ws.cell(row=row_start, column=7, value=mov.m_observaciones or '')
        
        # Aplicar estilos a los datos
        for col in range(1, 8):
            cell = ws.cell(row=row_start, column=col)
            cell.font = estilos['data_font']
            cell.border = estilos['data_border']
        
        row_start += 1
    
    # Ajustar ancho de columnas
    ajustar_columnas_excel(ws)
    
    # Crear respuesta
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    response = make_response(output.read())
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Content-Disposition'] = f'attachment; filename=CNM_Detalle_{item.i_codigo}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    
    return response

def _exportar_articulos_pdf(articulo, item, movimientos, saldo_calculado):
    """Exporta detalle de artículo a PDF con cabecera institucional"""
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4),
                          rightMargin=0.5*inch, leftMargin=0.5*inch,
                          topMargin=0.5*inch, bottomMargin=0.5*inch)
    
    elements = []
    
    # Crear cabecera institucional
    elements.extend(crear_cabecera_pdf(f"Detalle de Artículo - {item.i_codigo}"))
    
    # Obtener estilos
    styles = crear_estilos_pdf()
    
    # Título de la sección de información
    elements.append(Paragraph("INFORMACIÓN DEL ARTÍCULO", styles['TituloSeccion']))
    elements.append(Spacer(1, 12))
    
    # Información del artículo
    info_data = [
        ['Código:', item.i_codigo, 'Stock Actual:', str(item.i_cantidad)],
        ['Nombre:', item.i_nombre, 'Saldo Calculado:', str(saldo_calculado)],
        ['Valor Unitario:', f"${float(item.i_vUnitario):.2f}", 'Valor Total:', f"${float(item.i_vTotal):.2f}"],
        ['Stock Mín/Máx:', f"{articulo.a_stockMin}/{articulo.a_stockMax}", 'Cuenta Contable:', articulo.a_c_contable]
    ]
    
    info_table = Table(info_data, colWidths=[1.2*inch, 2.5*inch, 1.2*inch, 1.5*inch])
    info_table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTNAME', (2, 0), (2, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('LEFTPADDING', (0, 0), (-1, -1), 6),
        ('RIGHTPADDING', (0, 0), (-1, -1), 6),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
    ]))
    
    elements.append(info_table)
    elements.append(Spacer(1, 20))
    
    # Título de movimientos
    elements.append(Paragraph("HISTORIAL DE MOVIMIENTOS", styles['TituloSeccion']))
    elements.append(Paragraph(f"Mostrando los últimos 20 movimientos", styles['Normal']))
    elements.append(Spacer(1, 12))
    
    # Tabla de movimientos
    mov_data = [['Fecha', 'Tipo', 'Cantidad', 'Valor Unit.', 'Valor Total', 'Observaciones']]
    
    for mov in movimientos[:20]:  # Limitar a 20 movimientos más recientes
        mov_data.append([
            mov.m_fecha.strftime('%d/%m/%Y'),
            mov.m_tipo.title(),
            str(mov.m_cantidad),
            f"${float(mov.m_valorUnitario):.2f}",
            f"${float(mov.m_valorTotal):.2f}",
            (mov.m_observaciones or '')[:30] + '...' if mov.m_observaciones and len(mov.m_observaciones) > 30 else (mov.m_observaciones or '')
        ])
    
    mov_table = Table(mov_data, colWidths=[0.8*inch, 0.8*inch, 0.8*inch, 0.8*inch, 0.8*inch, 2.5*inch])
    
    # Aplicar estilo estandarizado
    aplicar_estilo_tabla_pdf(mov_table)
    
    elements.append(mov_table)
    
    # Construir PDF
    doc.build(elements)
    buffer.seek(0)
    
    response = make_response(buffer.read())
    response.headers['Content-Type'] = 'application/pdf'
    response.headers['Content-Disposition'] = f'attachment; filename=CNM_Detalle_{item.i_codigo}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf'
    
    return response

@bp.route('/<int:articulo_id>/salida', methods=['POST'])
@login_required
def registrar_salida(articulo_id):
    """Registra una salida de artículo con asignación a personal"""
    try:
        cantidad = int(request.form['cantidad'])
        valor_unitario = float(request.form['valor_unitario'])
        observaciones = request.form.get('observaciones', '')
        persona_id = request.form.get('persona_id')
        usuario_id = current_user.id
        
        # Validar que se haya seleccionado una persona
        if not persona_id:
            flash('Debe seleccionar una persona para asignar el artículo', 'error')
            return redirect(request.referrer or url_for('articulos.listar_articulos'))
        
        persona_id = int(persona_id)
        
        # Registrar la salida con asignación a personal
        articulo_service.registrar_salida_con_asignacion(
            articulo_id, cantidad, valor_unitario, usuario_id, persona_id, observaciones
        )
        
        flash('Artículo asignado exitosamente al personal', 'success')
    except ValueError as e:
        flash(str(e), 'error')
    except Exception as e:
        flash(f'Error al asignar artículo: {str(e)}', 'error')
    
    return redirect(request.referrer or url_for('articulos.listar_articulos'))

@bp.route('/exportar_asignaciones')
@login_required
def exportar_asignaciones():
    """Exportar historial de asignaciones a Excel o PDF"""
    export_format = request.args.get('export_format')
    
    # Obtener los filtros de la URL
    estado = request.args.get('estado', '').strip()
    articulo = request.args.get('articulo', '').strip()
    persona = request.args.get('persona', '').strip()

    # Construir la consulta base
    query = db.session.query(Consumo, Persona, Item).join(Persona, Consumo.pe_id == Persona.id).join(Item, Consumo.i_id == Item.id)

    # Aplicar filtros
    if estado:
        query = query.filter(Consumo.c_estado == estado)
    
    if articulo and articulo.lower() != "none":
        query = query.filter(
            db.or_(
                Item.i_nombre.ilike(f'%{articulo}%'),
                Item.i_codigo.ilike(f'%{articulo}%')
            )
        )
    
    if persona and persona.lower() != "none":
        query = query.filter(
            db.or_(
                Persona.pe_nombre.ilike(f'%{persona}%'),
                Persona.pe_apellido.ilike(f'%{persona}%')
            )
        )

    # Obtener los resultados con los filtros aplicados
    asignaciones = query.order_by(Consumo.c_fecha.desc(), Consumo.c_hora.desc()).all()

    # Manejo de exportación a Excel o PDF
    if export_format == 'excel':
        return _exportar_asignaciones_excel(asignaciones)
    elif export_format == 'pdf':
        return _exportar_asignaciones_pdf(asignaciones)
    else:
        flash("Formato de exportación no soportado.", "error")
        return redirect(url_for('articulos.listar_asignaciones'))


def _exportar_asignaciones_excel(asignaciones):
    """Exporta el historial de asignaciones a Excel con formato detallado y análisis estadístico"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Asignaciones Detallado"
    
    # Configurar orientación horizontal
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    
    # Crear cabecera institucional
    current_row = crear_cabecera_excel(ws, "Reporte Detallado de Asignaciones de Artículos")
    
    # Preparar datos detallados con estadísticas
    headers = [
        'Fecha', 'Hora', 'Artículo', 'Código', 'Personal Asignado', 'Cantidad',
        'Valor Unitario', 'Valor Total', 'Estado', 'Días Transcurridos',
        'Fecha Devolución', 'Observaciones'
    ]
    
    data = []
    valor_total_asignaciones = 0
    asignaciones_por_estado = {}
    asignaciones_por_persona = {}
    total_cantidad = 0
    
    for consumo, persona, item in asignaciones:
        # Calcular días transcurridos
        from datetime import datetime, date
        if isinstance(consumo.c_fecha, date):
            fecha_asignacion = consumo.c_fecha
        else:
            fecha_asignacion = consumo.c_fecha.date()
        
        dias_transcurridos = (datetime.now().date() - fecha_asignacion).days
        
        # Estadísticas
        valor_total_asignaciones += float(consumo.c_valorTotal)
        total_cantidad += consumo.c_cantidad
        
        # Contar por estado
        estado = consumo.c_estado
        asignaciones_por_estado[estado] = asignaciones_por_estado.get(estado, 0) + 1
        
        # Contar por persona
        nombre_completo = f"{persona.pe_nombre} {persona.pe_apellido}"
        asignaciones_por_persona[nombre_completo] = asignaciones_por_persona.get(nombre_completo, 0) + 1
        
        data.append([
            formatear_fecha(consumo.c_fecha),
            consumo.c_hora.strftime('%H:%M:%S') if hasattr(consumo, 'c_hora') and consumo.c_hora else 'N/A',
            item.i_nombre,
            item.i_codigo,
            nombre_completo,
            consumo.c_cantidad,
            formatear_valor_moneda(consumo.c_valorUnitario),
            formatear_valor_moneda(consumo.c_valorTotal),
            estado,
            dias_transcurridos,
            formatear_fecha(consumo.c_fecha_devolucion) if hasattr(consumo, 'c_fecha_devolucion') and consumo.c_fecha_devolucion else 'N/A',
            truncar_texto(consumo.c_observaciones or 'Sin observaciones', 40)
        ])
    
    # Crear tabla detallada
    current_row = crear_tabla_detallada_excel(ws, headers, data, current_row, "HISTORIAL DETALLADO DE ASIGNACIONES")
    
    # Agregar análisis estadístico completo
    current_row += 2
    ws[f'A{current_row}'] = "ANÁLISIS ESTADÍSTICO DE ASIGNACIONES"
    ws[f'A{current_row}'].font = Font(name='Calibri', size=14, bold=True, color="2E5984")
    ws.merge_cells(f'A{current_row}:F{current_row}')
    current_row += 2
    
    resumen_headers = ['Métrica', 'Cantidad', 'Valor', 'Porcentaje', 'Observaciones']
    resumen_data = [
        ['Total de Asignaciones', len(asignaciones), formatear_valor_moneda(valor_total_asignaciones), '100%', 'Asignaciones totales registradas'],
        ['Cantidad Total Asignada', total_cantidad, 'N/A', '100%', 'Unidades totales asignadas'],
        ['Valor Promedio por Asignación', f"{valor_total_asignaciones/len(asignaciones):.2f}" if asignaciones else "0", formatear_valor_moneda(valor_total_asignaciones/len(asignaciones)) if asignaciones else "$0.00", 'N/A', 'Valor promedio por asignación'],
        ['Personal Único Involucrado', len(asignaciones_por_persona), 'N/A', f"{(len(asignaciones_por_persona)/len(asignaciones)*100):.1f}%" if asignaciones else "0%", 'Personas diferentes con asignaciones']
    ]
    
    # Agregar desglose por estado
    for estado, cantidad in asignaciones_por_estado.items():
        porcentaje = f"{(cantidad/len(asignaciones)*100):.1f}%" if asignaciones else "0%"
        resumen_data.append([f'Asignaciones "{estado}"', cantidad, 'N/A', porcentaje, f'Estado: {estado}'])
    
    current_row = crear_tabla_detallada_excel(ws, resumen_headers, resumen_data, current_row)
    
    # Agregar ranking de personal con más asignaciones
    if asignaciones_por_persona:
        current_row += 2
        ws[f'A{current_row}'] = "RANKING DE PERSONAL CON MÁS ASIGNACIONES"
        ws[f'A{current_row}'].font = Font(name='Calibri', size=12, bold=True, color="2E5984")
        ws.merge_cells(f'A{current_row}:D{current_row}')
        current_row += 2
        
        ranking_headers = ['Personal', 'Cantidad de Asignaciones', 'Porcentaje', 'Observaciones']
        ranking_data = []
        
        # Ordenar por cantidad de asignaciones (top 10)
        top_personal = sorted(asignaciones_por_persona.items(), key=lambda x: x[1], reverse=True)[:10]
        
        for nombre, cantidad in top_personal:
            porcentaje = f"{(cantidad/len(asignaciones)*100):.1f}%" if asignaciones else "0%"
            ranking_data.append([
                nombre,
                cantidad,
                porcentaje,
                'Personal activo' if cantidad > 1 else 'Asignación única'
            ])
        
        current_row = crear_tabla_detallada_excel(ws, ranking_headers, ranking_data, current_row)
    
    # Ajustar ancho de columnas
    ajustar_columnas_excel(ws)
    
    # Crear respuesta
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    response = make_response(output.read())
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Content-Disposition'] = f'attachment; filename=CNM_Asignaciones_Detallado_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    
    return response


def _exportar_asignaciones_pdf(asignaciones):
    """Exporta el historial de asignaciones a PDF con formato detallado y análisis estadístico"""
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4),
                          rightMargin=0.5*inch, leftMargin=0.5*inch,
                          topMargin=0.5*inch, bottomMargin=0.5*inch)

    elements = []
    
    # Crear cabecera institucional
    elements.extend(crear_cabecera_pdf("Reporte Detallado de Asignaciones de Artículos"))
    
    # Calcular estadísticas
    valor_total_asignaciones = sum(float(consumo.c_valorTotal) for consumo, _, _ in asignaciones)
    total_cantidad = sum(consumo.c_cantidad for consumo, _, _ in asignaciones)
    asignaciones_por_estado = {}
    
    for consumo, _, _ in asignaciones:
        estado = consumo.c_estado
        asignaciones_por_estado[estado] = asignaciones_por_estado.get(estado, 0) + 1
    
    # Resumen estadístico
    resumen_headers = ['Métrica', 'Cantidad', 'Valor', 'Observaciones']
    resumen_data = [
        ['Total Asignaciones', str(len(asignaciones)), formatear_valor_moneda(valor_total_asignaciones), 'Asignaciones registradas'],
        ['Cantidad Total', str(total_cantidad), 'N/A', 'Unidades asignadas'],
        ['Valor Promedio', f"{valor_total_asignaciones/len(asignaciones):.2f}" if asignaciones else "0", formatear_valor_moneda(valor_total_asignaciones/len(asignaciones)) if asignaciones else "$0.00", 'Por asignación']
    ]
    
    # Agregar estados
    for estado, cantidad in asignaciones_por_estado.items():
        porcentaje = f"({(cantidad/len(asignaciones)*100):.1f}%)" if asignaciones else "(0%)"
        resumen_data.append([f'Estado "{estado}"', str(cantidad), porcentaje, f'Asignaciones en {estado.lower()}'])
    
    elements.extend(crear_tabla_detallada_pdf(
        resumen_headers, resumen_data,
        "RESUMEN ESTADÍSTICO DE ASIGNACIONES",
        [1.5*inch, 1*inch, 1.2*inch, 2*inch]
    ))

    # Datos detallados de la tabla
    data_headers = ['Fecha', 'Artículo', 'Personal', 'Cant.', 'V.Unit', 'V.Total', 'Estado', 'Días']
    data = []
    
    for consumo, persona, item in asignaciones:
        # Calcular días transcurridos
        from datetime import datetime, date
        if isinstance(consumo.c_fecha, date):
            fecha_asignacion = consumo.c_fecha
        else:
            fecha_asignacion = consumo.c_fecha.date()
        
        dias_transcurridos = (datetime.now().date() - fecha_asignacion).days
        
        data.append([
            formatear_fecha(consumo.c_fecha),
            truncar_texto(item.i_nombre, 15),
            truncar_texto(f"{persona.pe_nombre} {persona.pe_apellido}", 18),
            str(consumo.c_cantidad),
            formatear_valor_moneda(consumo.c_valorUnitario),
            formatear_valor_moneda(consumo.c_valorTotal),
            consumo.c_estado,
            str(dias_transcurridos)
        ])

    # Crear tabla detallada de asignaciones
    elements.extend(crear_tabla_detallada_pdf(
        data_headers, data,
        "HISTORIAL DETALLADO DE ASIGNACIONES",
        [0.8*inch, 1.5*inch, 1.5*inch, 0.5*inch, 0.7*inch, 0.8*inch, 0.8*inch, 0.5*inch]
    ))

    # Construir PDF
    doc.build(elements)
    buffer.seek(0)
    
    response = make_response(buffer.read())
    response.headers['Content-Type'] = 'application/pdf'
    response.headers['Content-Disposition'] = f'attachment; filename=CNM_Asignaciones_Detallado_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf'
    
    return response


@bp.route('/asignacion-multiple', methods=['GET', 'POST'])
@login_required
def asignacion_multiple():
    """Permite asignar múltiples artículos a una persona"""
    from app.database.models import Persona, Articulo, Item
    from app.services.articulo_service import ArticuloService
    
    articulo_service = ArticuloService()

    if request.method == 'POST':
        try:
            # Validar que se haya seleccionado una persona
            persona_id = request.form.get('persona_id')
            if not persona_id:
                flash('Debe seleccionar una persona', 'error')
                return redirect(url_for('articulos.asignacion_multiple'))
            
            persona_id = int(persona_id)
            
            # Verificar que la persona existe y está activa
            persona = Persona.query.get(persona_id)
            if not persona:
                flash('La persona seleccionada no existe', 'error')
                return redirect(url_for('articulos.asignacion_multiple'))
            
            if persona.pe_estado != 'Activo':
                flash('Solo se puede asignar artículos a personal activo', 'error')
                return redirect(url_for('articulos.asignacion_multiple'))
            
            articulos_data = []
            articulos_procesados = set()
            
            # Procesar múltiples artículos del formulario
            i = 0
            while f'articulo_id_{i}' in request.form:
                articulo_id = request.form.get(f'articulo_id_{i}')
                cantidad = request.form.get(f'cantidad_{i}')
                valor_unitario = request.form.get(f'valor_unitario_{i}')
                
                if articulo_id and cantidad and valor_unitario:
                    try:
                        articulo_id = int(articulo_id)
                        cantidad = int(cantidad)
                        valor_unitario = float(valor_unitario)
                        
                        # Validaciones básicas
                        if cantidad <= 0:
                            flash('La cantidad debe ser mayor a 0', 'error')
                            return redirect(url_for('articulos.asignacion_multiple'))
                        
                        if valor_unitario <= 0:
                            flash('El valor unitario debe ser mayor a 0', 'error')
                            return redirect(url_for('articulos.asignacion_multiple'))
                        
                        # Verificar duplicados
                        if articulo_id in articulos_procesados:
                            flash('No puede seleccionar el mismo artículo múltiples veces', 'error')
                            return redirect(url_for('articulos.asignacion_multiple'))
                        
                        # Verificar que el artículo existe y tiene stock
                        resultado = articulo_service.obtener_por_id(articulo_id)
                        if not resultado:
                            flash(f'Artículo con ID {articulo_id} no encontrado', 'error')
                            return redirect(url_for('articulos.asignacion_multiple'))
                        
                        articulo, item = resultado
                        
                        # Verificar stock disponible
                        if item.i_cantidad < cantidad:
                            flash(f'Stock insuficiente para {item.i_nombre}. Disponible: {item.i_cantidad}, Solicitado: {cantidad}', 'error')
                            return redirect(url_for('articulos.asignacion_multiple'))
                        
                        articulos_data.append({
                            'articulo_id': articulo_id,
                            'cantidad': cantidad,
                            'valor_unitario': valor_unitario,
                            'nombre': item.i_nombre,
                            'codigo': item.i_codigo
                        })
                        articulos_procesados.add(articulo_id)
                        
                    except (ValueError, TypeError) as e:
                        flash(f'Error en los datos del artículo {i+1}: valores inválidos', 'error')
                        return redirect(url_for('articulos.asignacion_multiple'))
                
                i += 1
            
            # Validar que se hayan agregado artículos
            if not articulos_data:
                flash('Debe agregar al menos un artículo válido', 'error')
                return redirect(url_for('articulos.asignacion_multiple'))
            
            observaciones = request.form.get('observaciones', '').strip()
            
            # Procesar todas las asignaciones
            asignaciones_exitosas = 0
            errores = []
            
            try:
                # Usar transacción para asegurar consistencia
                for articulo_data in articulos_data:
                    try:
                        # Registrar la asignación
                        movimiento, consumo = articulo_service.registrar_salida_con_asignacion(
                            articulo_data['articulo_id'],
                            articulo_data['cantidad'],
                            articulo_data['valor_unitario'],
                            current_user.id,
                            persona_id,
                            f"Asignación múltiple - {observaciones}" if observaciones else "Asignación múltiple"
                        )
                        asignaciones_exitosas += 1
                        
                    except Exception as e:
                        errores.append(f"{articulo_data['codigo']} - {articulo_data['nombre']}: {str(e)}")
                        db.session.rollback()
                        continue
                
                # Mostrar resultados
                if asignaciones_exitosas > 0:
                    flash(f'Se procesaron {asignaciones_exitosas} asignación(es) exitosamente para {persona.pe_nombre}', 'success')
                
                if errores:
                    for error in errores:
                        flash(f'Error: {error}', 'warning')
                
                # Redirigir según el resultado
                if asignaciones_exitosas > 0:
                    # Redirigir a vista previa con datos de las asignaciones procesadas
                    articulos_procesados_ids = [str(data['articulo_id']) for data in articulos_data]
                    return redirect(url_for('articulos.vista_previa_asignaciones',
                                          persona_id=persona_id,
                                          articulos=','.join(articulos_procesados_ids),
                                          observaciones=observaciones))
                else:
                    return redirect(url_for('articulos.asignacion_multiple'))
                    
            except Exception as e:
                db.session.rollback()
                flash(f'Error general en el procesamiento: {str(e)}', 'error')
                return redirect(url_for('articulos.asignacion_multiple'))
            
        except Exception as e:
            flash(f'Error en la asignación múltiple: {str(e)}', 'error')
            return redirect(url_for('articulos.asignacion_multiple'))
    
    # GET: Mostrar formulario
    try:
        # Obtener personas activas
        personas = Persona.query.filter_by(pe_estado='Activo').all()
        
        # Obtener artículos con stock disponible
        articulos = db.session.query(Articulo, Item).join(
            Item, Articulo.i_id == Item.id
        ).filter(
            Item.i_cantidad > 0
        ).order_by(Item.i_nombre).all()
        
        return render_template('articulos/asignacion_multiple.html',
                             personas=personas,
                             articulos=articulos)
                             
    except Exception as e:
        flash(f'Error al cargar el formulario: {str(e)}', 'error')
        return redirect(url_for('articulos.listar_asignaciones'))

@bp.route('/vista-previa-asignaciones')
@login_required
def vista_previa_asignaciones():
    """Vista previa informativa de asignaciones procesadas con opción de exportación PDF"""
    from app.database.models import Persona, Consumo, Item
    
    try:
        # Obtener parámetros
        persona_id = request.args.get('persona_id', type=int)
        articulos_ids = request.args.get('articulos', '').split(',')
        observaciones = request.args.get('observaciones', '')
        export_format = request.args.get('export')
        
        if not persona_id or not articulos_ids or not articulos_ids[0]:
            flash('Datos de vista previa no válidos', 'error')
            return redirect(url_for('articulos.listar_asignaciones'))
        
        # Obtener información de la persona
        persona = Persona.query.get(persona_id)
        if not persona:
            flash('Persona no encontrada', 'error')
            return redirect(url_for('articulos.listar_asignaciones'))
        
        # Obtener las asignaciones más recientes de esta persona para los artículos especificados
        asignaciones_query = db.session.query(Consumo, Item).join(
            Item, Consumo.i_id == Item.id
        ).filter(
            Consumo.pe_id == persona_id,
            Consumo.i_id.in_([int(aid) for aid in articulos_ids if aid.isdigit()])
        ).order_by(Consumo.c_fecha.desc(), Consumo.c_hora.desc())
        
        # Limitar a las asignaciones más recientes (últimas 10 por seguridad)
        asignaciones_recientes = asignaciones_query.limit(len(articulos_ids)).all()
        
        # Calcular totales
        total_articulos = len(asignaciones_recientes)
        total_cantidad = sum(consumo.c_cantidad for consumo, _ in asignaciones_recientes)
        total_valor = sum(float(consumo.c_valorTotal) for consumo, _ in asignaciones_recientes)
        
        # Manejar exportación PDF
        if export_format == 'pdf':
            return _exportar_vista_previa_pdf(persona, asignaciones_recientes, observaciones,
                                            total_articulos, total_cantidad, total_valor)
        
        return render_template('articulos/vista_previa_asignaciones.html',
                             persona=persona,
                             asignaciones=asignaciones_recientes,
                             observaciones=observaciones,
                             total_articulos=total_articulos,
                             total_cantidad=total_cantidad,
                             total_valor=total_valor,
                             fecha_asignacion=datetime.now().strftime('%d/%m/%Y %H:%M:%S'))
                             
    except Exception as e:
        flash(f'Error al cargar vista previa: {str(e)}', 'error')
        return redirect(url_for('articulos.listar_asignaciones'))

def _exportar_vista_previa_pdf(persona, asignaciones, observaciones, total_articulos, total_cantidad, total_valor):
    """Exporta la vista previa de asignaciones a PDF"""
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4,
                          rightMargin=0.75*inch, leftMargin=0.75*inch,
                          topMargin=0.75*inch, bottomMargin=0.75*inch)
    
    elements = []
    
    # Crear cabecera institucional
    elements.extend(crear_cabecera_pdf("Vista Previa de Asignaciones Múltiples"))
    
    # Información de la persona asignada
    elements.append(Paragraph("INFORMACIÓN DEL PERSONAL", crear_estilos_pdf()['TituloSeccion']))
    
    persona_data = [
        ['Nombre Completo:', f"{persona.pe_nombre} {persona.pe_apellido or ''}"],
        ['Cargo:', persona.pe_cargo or 'No especificado'],
        ['Estado:', persona.pe_estado],
        ['Fecha de Asignación:', datetime.now().strftime('%d/%m/%Y %H:%M:%S')]
    ]
    
    if observaciones:
        persona_data.append(['Observaciones:', observaciones])
    
    persona_table = Table(persona_data, colWidths=[2*inch, 4*inch])
    persona_table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTNAME', (1, 0), (1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#F8F9FA'))
    ]))
    
    elements.append(persona_table)
    elements.append(Spacer(1, 20))
    
    # Detalle de artículos asignados
    detalle_headers = ['Artículo', 'Código', 'Cantidad', 'Valor Unit.', 'Valor Total', 'Estado', 'Fecha']
    detalle_data = []
    
    for consumo, item in asignaciones:
        detalle_data.append([
            truncar_texto(item.i_nombre, 25),
            item.i_codigo,
            str(consumo.c_cantidad),
            formatear_valor_moneda(consumo.c_valorUnitario),
            formatear_valor_moneda(consumo.c_valorTotal),
            consumo.c_estado,
            formatear_fecha(consumo.c_fecha)
        ])
    
    elements.extend(crear_tabla_detallada_pdf(
        detalle_headers, detalle_data,
        "DETALLE DE ARTÍCULOS ASIGNADOS",
        [1.8*inch, 0.8*inch, 0.6*inch, 0.8*inch, 0.8*inch, 0.8*inch, 0.8*inch]
    ))
    
    # Apartado de firmas
    elements.append(Spacer(1, 30))
    elements.append(Paragraph("FIRMAS DE RESPONSABILIDAD", crear_estilos_pdf()['TituloSeccion']))
    elements.append(Spacer(1, 20))
    
    # Obtener nombre del usuario actual (quien despacha)
    nombre_despachador = current_user.username if current_user.is_authenticated else 'Administrador'
    
    # Tabla de firmas simplificada (2x4)
    firmas_data = [
        ['ENTREGA', 'RECIBE'],
        [f'Nombre: {nombre_despachador}', f'Nombre: {persona.pe_nombre} {persona.pe_apellido or ""}'],
        [f'Cargo: Administrador del Sistema', f'Cargo: {persona.pe_cargo or "No especificado"}'],
        ['Firma: _________________________', 'Firma: _________________________']
    ]
    
    firmas_table = Table(firmas_data, colWidths=[3*inch, 3*inch])
    firmas_table.setStyle(TableStyle([
        # Encabezados
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2E5984')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        
        # Información de nombres y cargos (filas 1-2)
        ('FONTNAME', (0, 1), (-1, 2), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, 2), 10),
        ('ALIGN', (0, 1), (-1, 2), 'CENTER'),
        ('VALIGN', (0, 1), (-1, 2), 'MIDDLE'),
        
        # Líneas de firma (fila 3)
        ('FONTNAME', (0, 3), (-1, 3), 'Helvetica'),
        ('FONTSIZE', (0, 3), (-1, 3), 10),
        ('ALIGN', (0, 3), (-1, 3), 'CENTER'),
        ('VALIGN', (0, 3), (-1, 3), 'MIDDLE'),
        
        # Bordes
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('LEFTPADDING', (0, 0), (-1, -1), 12),
        ('RIGHTPADDING', (0, 0), (-1, -1), 12),
        ('TOPPADDING', (0, 3), (-1, 3), 30),
        ('BOTTOMPADDING', (0, 3), (-1, 3), 30),
    ]))
    
    elements.append(firmas_table)
    elements.append(Spacer(1, 20))
    
    # Nota final
    nota_style = ParagraphStyle(
        'NotaFinal',
        parent=crear_estilos_pdf()['Normal'],
        fontSize=8,
        fontName='Helvetica-Oblique',
        textColor=colors.HexColor('#666666'),
        alignment=1
    )
    elements.append(Paragraph("Este documento certifica la entrega y recepción de los artículos detallados anteriormente.", nota_style))
    
    # Construir PDF
    doc.build(elements)
    buffer.seek(0)
    
    response = make_response(buffer.read())
    response.headers['Content-Type'] = 'application/pdf'
    response.headers['Content-Disposition'] = f'attachment; filename=CNM_Vista_Previa_Asignaciones_{persona.pe_nombre.replace(" ", "_")}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf'
    
    return response
